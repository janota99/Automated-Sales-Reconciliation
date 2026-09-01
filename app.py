"""Professional QuickBooks-to-Infinium sales reconciliation.

Run with:
    python -m streamlit run app.py

This module is the application layer only: Streamlit UI, Excel workpaper
construction, and top-level orchestration. Source ingestion (reading files,
header detection, cleanup, column mapping, and pre-match validation) lives
in ``ingestion.py``. Every normalization, matching, and analytical rule
lives in ``matching.py``. Keeping those concerns out of this file is what
keeps app.py short and focused on "how results are shown," not "how they
are computed."

The app produces two files from one controlled reconciliation run:
    1. Sales_Reconciliation_<run>.xlsx
       Raw Data, Reconciled Data, Unresolved Exceptions, and
       Product Aggregate Summary.
    2. Sales_Reconciliation_Analytics_<run>.xlsx
       Optional technical evidence, normalization, method analytics,
       controls, and run configuration.
"""

from __future__ import annotations

import hashlib
import html
import io
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Optional
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ingestion import (
    build_source_validation_report,
    detect_header_row,
    file_sha256,
    filter_qb_subtotal_rows,
    fiscal_period_column_profile,
    list_source_sheets,
    mapping_panel,
    read_source_file,
    secondary_mapping_panel,
)
from matching import (
    AMOUNT_CENTS,
    APP_VERSION,
    INF_ID,
    MATCHING_RULE_VERSION,
    QB_ID,
    ReconciliationResult,
    build_fiscal_exception_summary,
    build_reconciliation,
    cents_or_zero,
    cents_to_float,
    numeric_quantity_sum,
    numeric_sum,
    valid_cents,
)
from assets.ui_assets import INFOR_LOGO_URI, QUICKBOOKS_LOGO_URI

CENTRAL_TIMEZONE = ZoneInfo("America/Chicago")

NAVY = "1B365D"
NAVY_LIGHT = "E8EEF5"
TEAL = "27666B"
TEAL_LIGHT = "E7F1F1"
SLATE = "475569"
SLATE_LIGHT = "F2F5F8"
AMBER = "FFF3CD"
ORANGE = "FCE8D5"
DUPLICATE_RED_FILL = "FFC7CE"
DUPLICATE_RED_TEXT = "9C0006"
RED_LIGHT = "FDECEC"
GREEN_LIGHT = "E8F3EC"
WHITE = "FFFFFF"
TEXT = "172B4D"
BORDER = "D5DDE4"
TOTAL_FILL = "E2E8F0"


def format_central_timestamp(value: datetime) -> str:
    """Format a timezone-aware report timestamp in U.S. Central Time."""
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc)
    central_value = value.astimezone(CENTRAL_TIMEZONE)
    return central_value.strftime("%Y-%m-%d %I:%M:%S %p %Z")


def excel_safe(value: Any) -> Any:
    """Convert values safely for Excel and neutralize formula-like source text."""
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(value, pd.Timestamp):
        value = value.to_pydatetime()
    if isinstance(value, datetime) and value.tzinfo is not None:
        value = value.replace(tzinfo=None)
    if isinstance(value, Decimal):
        return float(value)
    if hasattr(value, "item"):
        try:
            value = value.item()
        except Exception:
            pass
    if isinstance(value, (list, tuple, set)):
        value = "; ".join(str(item) for item in value)
    if isinstance(value, str):
        # Excel treats strings beginning with these characters as formulas or
        # external commands in several import/opening contexts. Prefixing an
        # apostrophe forces a literal text cell without changing its display.
        if value.lstrip().startswith(("=", "+", "-", "@")):
            return "'" + value
    return value


# ---------------------------------------------------------------------------
# Excel construction
# ---------------------------------------------------------------------------


def _thin_border() -> Border:
    side = Side(style="thin", color=BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _total_border() -> Border:
    return Border(
        top=Side(style="thin", color=SLATE),
        bottom=Side(style="double", color=SLATE),
    )


def _write_dataframe_values(ws, frame: pd.DataFrame, start_row: int, start_col: int) -> None:
    for col_offset, header in enumerate(frame.columns):
        ws.cell(start_row, start_col + col_offset, excel_safe(str(header)))
    for row_offset, row in enumerate(frame.itertuples(index=False, name=None), 1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset, excel_safe(value))


def _format_header(ws, row: int, start_col: int, end_col: int, color: str) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=color))
    ws.row_dimensions[row].height = 34


def _format_body_block(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    light_fill: str,
) -> None:
    border = _thin_border()
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=light_fill) if row % 2 == 0 else PatternFill(fill_type=None)
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="Segoe UI", size=10, color=TEXT)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20


def _apply_duplicate_style(ws, row: int, start_col: int, end_col: int) -> None:
    """Apply Excel's traditional red bad-value style to a duplicate source row."""
    fill = PatternFill("solid", fgColor=DUPLICATE_RED_FILL)
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = fill
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=DUPLICATE_RED_TEXT)


def _apply_number_formats(
    ws,
    headers: list[str],
    start_row: int,
    end_row: int,
    start_col: int,
    amount_columns: set[str],
    quantity_columns: set[str],
) -> None:
    for offset, header in enumerate(headers):
        col = start_col + offset
        header_upper = str(header).upper()
        for row in range(start_row, end_row + 1):
            cell = ws.cell(row, col)
            if any(term in header_upper for term in ("VALID", "GROUP-LEVEL", "AUTOMATIC")):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif any(term in header_upper for term in ("PERCENT", "RATE", "SHARE")):
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif any(term in header_upper for term in ("COUNT", "ROWS", "CANDIDATE COUNT")):
                cell.number_format = '#,##0;[Red](#,##0);-'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif header in amount_columns or any(term in header_upper for term in ("AMOUNT", "VALUE", "VARIANCE", "DIFFERENCE", "BALANCE", "EXPOSURE", "TOLERANCE")):
                cell.number_format = '$#,##0.00;[Red]($#,##0.00);-'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif header in quantity_columns or any(term in header_upper for term in ("QUANTITY", "QTY", "COUNT", "ROWS")):
                cell.number_format = '#,##0.00;[Red](#,##0.00);-'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif "DATE" in header_upper or "TIMESTAMP" in header_upper:
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="center", vertical="center")


def _set_widths(
    ws,
    start_col: int,
    end_col: int,
    start_row: int,
    end_row: int,
    minimum: float = 11,
    maximum: float = 34,
) -> None:
    for col in range(start_col, end_col + 1):
        lengths = [len(str(ws.cell(row, col).value or "")) for row in range(start_row, end_row + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max(lengths, default=0) + 2, minimum), maximum)


def _autofit_workbook_columns(
    wb: Workbook,
    minimum: float = 10,
    maximum: float = 52,
    sample_rows: int = 750,
) -> None:
    """Auto-size columns from a bounded sample instead of rescanning every cell."""
    for ws in wb.worksheets:
        merged_coordinates: set[str] = set()
        for merged_range in ws.merged_cells.ranges:
            for row in ws.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                merged_coordinates.update(cell.coordinate for cell in row)

        sampled_row_numbers = list(range(1, min(ws.max_row, sample_rows) + 1))
        if ws.max_row > sample_rows:
            sampled_row_numbers.append(ws.max_row)

        for column_index in range(1, ws.max_column + 1):
            letter = get_column_letter(column_index)
            maximum_length = 0
            has_unmerged_value = False
            for row_index in sampled_row_numbers:
                cell = ws.cell(row_index, column_index)
                if cell.coordinate in merged_coordinates or cell.value is None:
                    continue
                has_unmerged_value = True
                value = cell.value
                if isinstance(value, datetime):
                    display = value.strftime("%Y-%m-%d %I:%M:%S %p")
                elif isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                    if "$" in str(cell.number_format):
                        display = f"${float(value):,.2f}"
                    elif any(token in str(cell.number_format) for token in ("0.00", "#,##0")):
                        display = f"{float(value):,.2f}"
                    else:
                        display = str(value)
                else:
                    display = str(value)
                maximum_length = max(
                    maximum_length,
                    max((len(line) for line in display.splitlines()), default=0),
                )

            if has_unmerged_value:
                ws.column_dimensions[letter].width = min(max(maximum_length + 2, minimum), maximum)
            else:
                existing = ws.column_dimensions[letter].width or minimum
                ws.column_dimensions[letter].width = min(max(existing, 3.5), maximum)


def _write_title_band(ws, row: int, start_col: int, end_col: int, title: str, color: str) -> None:
    if start_col < end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, title)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.font = Font(name="Segoe UI", size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=color)
    ws.row_dimensions[row].height = 27


def _write_caption_band(ws, row: int, start_col: int, end_col: int, caption: str, color: str) -> None:
    if start_col < end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, caption)
    cell.fill = PatternFill("solid", fgColor=SLATE_LIGHT)
    cell.font = Font(name="Segoe UI", size=9, italic=True, color=color)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row, col).fill = PatternFill("solid", fgColor=SLATE_LIGHT)
    ws.row_dimensions[row].height = 30


def _write_total_row(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    totals: dict[str, float],
    headers: list[str],
    label: str = "TOTAL",
) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=TOTAL_FILL)
        cell.font = Font(name="Segoe UI", size=10, bold=True, color=TEXT)
        cell.border = _total_border()
    ws.cell(row, start_col, label)
    for offset, header in enumerate(headers):
        if header in totals:
            cell = ws.cell(row, start_col + offset, totals[header])
            cell.number_format = '$#,##0.00;[Red]($#,##0.00);-' if "amount" in header.lower() or "value" in header.lower() else '#,##0.00;[Red](#,##0.00);-'
            cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[row].height = 23


def _source_totals(frame: pd.DataFrame, mapping: dict[str, Optional[str]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    amount_col = mapping.get("amount")
    quantity_col = mapping.get("quantity")
    if amount_col:
        totals[amount_col] = numeric_sum(frame[amount_col])
    if quantity_col:
        totals[quantity_col] = numeric_quantity_sum(frame[quantity_col])
    return totals


def _duplicate_source_indexes(
    result: ReconciliationResult,
    dataset: str,
) -> set[int]:
    """Return duplicate primary-row indexes, including for pre-2.9 session results."""
    attribute = "duplicate_qb_rows" if dataset == "QuickBooks" else "duplicate_inf_rows"
    stored_indexes = getattr(result, attribute, None)
    if stored_indexes is not None:
        return {int(index) for index in stored_indexes}

    analysis = getattr(result, "duplicate_analysis", pd.DataFrame())
    if analysis.empty or not {"Dataset", "Source Row IDs"}.issubset(analysis.columns):
        return set()
    duplicate_ids: set[str] = set()
    source_rows = analysis.loc[analysis["Dataset"].eq(dataset), "Source Row IDs"]
    for value in source_rows.dropna().astype(str):
        duplicate_ids.update(item.strip() for item in value.split(";") if item.strip())
    frame = result.qb_work if dataset == "QuickBooks" else result.inf_work
    id_column = QB_ID if dataset == "QuickBooks" else INF_ID
    return {
        int(index)
        for index in frame.index
        if str(frame.at[index, id_column]) in duplicate_ids
    }


def _resolved_paired_record(
    result: ReconciliationResult,
    record: dict[str, Any],
) -> dict[str, Any]:
    """Upgrade a paired-row record retained in Streamlit session state."""
    resolved = dict(record)
    if "QB Record Scope" in resolved and "Infinium Record Scope" in resolved:
        return resolved

    resolved["QB Record Scope"] = "Primary" if resolved.get("QB Index") is not None else None
    resolved["Infinium Record Scope"] = (
        "Primary" if resolved.get("Infinium Index") is not None else None
    )
    if resolved.get("Section") != "01 Matched - Historical Clearance":
        return resolved

    clearances = getattr(result, "historical_clearances", pd.DataFrame())
    if clearances.empty or "Clearance ID" not in clearances.columns:
        return resolved
    selected = clearances.loc[clearances["Clearance ID"].eq(resolved.get("Match ID"))]
    if selected.empty:
        return resolved
    clearance = selected.iloc[0]
    primary_is_qb = clearance["Primary Dataset"] == "QuickBooks Primary"
    resolved["QB Index"] = int(
        clearance["Primary Row Index"] if primary_is_qb else clearance["Secondary Row Index"]
    )
    resolved["Infinium Index"] = int(
        clearance["Secondary Row Index"] if primary_is_qb else clearance["Primary Row Index"]
    )
    resolved["QB Record Scope"] = "Primary" if primary_is_qb else "Historical"
    resolved["Infinium Record Scope"] = "Historical" if primary_is_qb else "Primary"
    return resolved


def _prepare_sheet(ws, landscape: bool = True) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if landscape else ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False


def build_raw_data_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.active
    ws.title = "Raw Data"
    qb_headers = list(result.qb_raw.columns)
    inf_headers = list(result.inf_raw.columns)
    qb_start = 1
    separator_col = len(qb_headers) + 1
    inf_start = separator_col + 1
    header_row = 3
    data_row = 4
    qb_end = len(qb_headers)
    inf_end = inf_start + len(inf_headers) - 1

    _write_title_band(ws, 1, qb_start, qb_end, "QUICKBOOKS | RAW TRANSACTION DETAIL", NAVY)
    _write_title_band(ws, 1, inf_start, inf_end, "INFINIUM | RAW UPLOAD", TEAL)
    _write_caption_band(
        ws, 2, qb_start, qb_end,
        f"{len(result.qb_raw):,} rows | Source control total: ${result.metrics['QuickBooks Source Total']:,.2f} | "
        f"{result.metrics['QuickBooks Subtotal Rows Excluded']:,} subtotal row(s) excluded before matching | "
        f"Generated {format_central_timestamp(result.run_timestamp)}",
        NAVY,
    )
    _write_caption_band(
        ws, 2, inf_start, inf_end,
        f"{len(result.inf_raw):,} rows | Source control total: ${result.metrics['Infinium Source Total']:,.2f} | "
        f"Values preserved before matching | Generated {format_central_timestamp(result.run_timestamp)}",
        TEAL,
    )
    _write_dataframe_values(ws, result.qb_raw, header_row, qb_start)
    _write_dataframe_values(ws, result.inf_raw, header_row, inf_start)
    _format_header(ws, header_row, qb_start, qb_end, NAVY)
    _format_header(ws, header_row, inf_start, inf_end, TEAL)
    _format_body_block(ws, data_row, data_row + len(result.qb_raw) - 1, qb_start, qb_end, NAVY_LIGHT)
    _format_body_block(ws, data_row, data_row + len(result.inf_raw) - 1, inf_start, inf_end, TEAL_LIGHT)
    for source_index in _duplicate_source_indexes(result, "QuickBooks"):
        if 0 <= source_index < len(result.qb_raw):
            _apply_duplicate_style(ws, data_row + source_index, qb_start, qb_end)
    for source_index in _duplicate_source_indexes(result, "Infinium"):
        if 0 <= source_index < len(result.inf_raw):
            _apply_duplicate_style(ws, data_row + source_index, inf_start, inf_end)
    qb_total_row = data_row + len(result.qb_raw)
    inf_total_row = data_row + len(result.inf_raw)
    _write_total_row(ws, qb_total_row, qb_start, qb_end, _source_totals(result.qb_raw, result.qb_mapping), qb_headers, "SOURCE TOTAL")
    _write_total_row(ws, inf_total_row, inf_start, inf_end, _source_totals(result.inf_raw, result.inf_mapping), inf_headers, "SOURCE TOTAL")
    _apply_number_formats(ws, qb_headers, data_row, qb_total_row, qb_start,
                          {result.qb_mapping["amount"]}, {result.qb_mapping.get("quantity") or ""})
    _apply_number_formats(ws, inf_headers, data_row, inf_total_row, inf_start,
                          {result.inf_mapping["amount"]}, set())
    ws.column_dimensions[get_column_letter(separator_col)].width = 3.5
    ws.column_dimensions[get_column_letter(separator_col)].fill = PatternFill("solid", fgColor=WHITE)
    _set_widths(ws, qb_start, qb_end, header_row, qb_total_row)
    _set_widths(ws, inf_start, inf_end, header_row, inf_total_row)
    ws.freeze_panes = f"{get_column_letter(inf_start)}{data_row}"
    ws.print_title_rows = "1:3"
    _prepare_sheet(ws)


def _paired_display_frames(result: ReconciliationResult) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    def ordered_union(primary_headers: list[str], historical_headers: list[str]) -> list[str]:
        output = list(primary_headers)
        output.extend(header for header in historical_headers if header not in output)
        return output

    def unique_context_header(base: str, headers: list[str]) -> str:
        candidate = base
        suffix = 2
        while candidate in headers:
            candidate = f"{base} {suffix}"
            suffix += 1
        return candidate

    paired_records = [
        _resolved_paired_record(result, record) for record in result.paired_rows
    ]
    qb_historical_used = any(
        record.get("QB Record Scope") == "Historical" for record in paired_records
    )
    inf_historical_used = any(
        record.get("Infinium Record Scope") == "Historical" for record in paired_records
    )
    qb_historical_headers = (
        list(result.qb_secondary_raw.columns)
        if qb_historical_used and result.qb_secondary_raw is not None else []
    )
    inf_historical_headers = (
        list(result.inf_secondary_raw.columns)
        if inf_historical_used and result.inf_secondary_raw is not None else []
    )
    qb_headers = ordered_union(list(result.qb_raw.columns), qb_historical_headers)
    inf_headers = ordered_union(list(result.inf_raw.columns), inf_historical_headers)
    qb_context_header = unique_context_header("QuickBooks Record Context", qb_headers)
    inf_context_header = unique_context_header("Infinium Record Context", inf_headers)

    def row_values(
        index: Optional[int],
        scope: Optional[str],
        primary_frame: pd.DataFrame,
        historical_frame: Optional[pd.DataFrame],
        headers: list[str],
    ) -> list[Any]:
        if index is None:
            return [None] * len(headers)
        frame = historical_frame if scope == "Historical" else primary_frame
        if frame is None:
            raise ValueError("A matched historical row is unavailable for display.")
        return [frame.at[index, header] if header in frame.columns else None for header in headers]

    qb_rows, inf_rows, match_results = [], [], []
    for record in paired_records:
        qidx, iidx = record["QB Index"], record["Infinium Index"]
        qb_scope = record.get("QB Record Scope")
        inf_scope = record.get("Infinium Record Scope")
        qb_values = row_values(
            qidx, qb_scope, result.qb_work, result.qb_secondary_work, qb_headers
        )
        inf_values = row_values(
            iidx, inf_scope, result.inf_work, result.inf_secondary_work, inf_headers
        )
        qb_values.append(
            "QuickBooks Prior Period Match"
            if qb_scope == "Historical" else "Primary QuickBooks Upload" if qidx is not None else None
        )
        inf_values.append(
            "Infinium Prior Period Match"
            if inf_scope == "Historical" else "Primary Infinium Upload" if iidx is not None else None
        )
        qb_rows.append(qb_values)
        inf_rows.append(inf_values)
        match_results.append(record["Match Result"])
    return (
        pd.DataFrame(qb_rows, columns=qb_headers + [qb_context_header]),
        pd.DataFrame(inf_rows, columns=inf_headers + [inf_context_header]),
        match_results,
    )


def build_reconciled_data_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Reconciled Data")
    qb_display, inf_display, match_results = _paired_display_frames(result)
    qb_headers = list(qb_display.columns)
    inf_headers = list(inf_display.columns)
    qb_start = 1
    match_col = len(qb_headers) + 1
    inf_start = match_col + 1
    qb_end = len(qb_headers)
    inf_end = inf_start + len(inf_headers) - 1
    header_row, data_row = 3, 4
    final_data_row = data_row + len(match_results) - 1

    _write_title_band(ws, 1, qb_start, qb_end, "QUICKBOOKS | RECONCILED", NAVY)
    _write_title_band(ws, 1, match_col, match_col, "MATCH RESULT", SLATE)
    _write_title_band(ws, 1, inf_start, inf_end, "INFINIUM | RECONCILED", TEAL)
    _write_caption_band(
        ws, 2, qb_start, qb_end,
        f"Every primary QuickBooks record appears once. Any accepted QuickBooks prior-period match is displayed on this side and labeled in Record Context. Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    _write_caption_band(ws, 2, match_col, match_col, "Matching Methodology", SLATE)
    _write_caption_band(
        ws, 2, inf_start, inf_end,
        "Every primary Infinium record appears once. Accepted prior-period matches are displayed; unused historical rows are excluded.",
        TEAL,
    )
    _write_dataframe_values(ws, qb_display, header_row, qb_start)
    ws.cell(header_row, match_col, "Match Result")
    for offset, value in enumerate(match_results, 1):
        ws.cell(header_row + offset, match_col, value)
    _write_dataframe_values(ws, inf_display, header_row, inf_start)
    _format_header(ws, header_row, qb_start, qb_end, NAVY)
    _format_header(ws, header_row, match_col, match_col, SLATE)
    _format_header(ws, header_row, inf_start, inf_end, TEAL)
    _format_body_block(ws, data_row, final_data_row, qb_start, qb_end, NAVY_LIGHT)
    _format_body_block(ws, data_row, final_data_row, match_col, match_col, SLATE_LIGHT)
    _format_body_block(ws, data_row, final_data_row, inf_start, inf_end, TEAL_LIGHT)

    duplicate_qb_rows = _duplicate_source_indexes(result, "QuickBooks")
    duplicate_inf_rows = _duplicate_source_indexes(result, "Infinium")
    for offset, stored_record in enumerate(result.paired_rows):
        record = _resolved_paired_record(result, stored_record)
        row = data_row + offset
        status = str(record["Section"])
        if status == "02 Unmatched QuickBooks":
            for col in range(qb_start, match_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=AMBER)
        elif status == "03 Unmatched Infinium":
            for col in range(match_col, inf_end + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=ORANGE)
        qidx = record["QB Index"]
        iidx = record["Infinium Index"]
        if (
            record.get("QB Record Scope") == "Primary"
            and qidx is not None
            and int(qidx) in duplicate_qb_rows
        ):
            _apply_duplicate_style(ws, row, qb_start, qb_end)
        if (
            record.get("Infinium Record Scope") == "Primary"
            and iidx is not None
            and int(iidx) in duplicate_inf_rows
        ):
            _apply_duplicate_style(ws, row, inf_start, inf_end)
        if record.get("QB Record Scope") == "Historical":
            ws.cell(row, qb_end).font = Font(
                name="Segoe UI", size=10, bold=True, color=NAVY
            )
        if record.get("Infinium Record Scope") == "Historical":
            ws.cell(row, inf_end).font = Font(
                name="Segoe UI", size=10, bold=True, color=TEAL
            )
        ws.cell(row, match_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30 if "group-level" in str(record["Match Result"]) else 21

    total_row = final_data_row + 1
    _write_total_row(ws, total_row, qb_start, qb_end,
                     _source_totals(result.qb_raw, result.qb_mapping), qb_headers, "RECONCILED TOTAL")
    _write_total_row(ws, total_row, inf_start, inf_end,
                     _source_totals(result.inf_raw, result.inf_mapping), inf_headers, "RECONCILED TOTAL")
    ws.cell(total_row, match_col, f"Control: {result.metrics['Control Status']}")
    ws.cell(total_row, match_col).fill = PatternFill("solid", fgColor=GREEN_LIGHT if result.metrics["Control Status"] == "PASS" else RED_LIGHT)
    ws.cell(total_row, match_col).font = Font(name="Segoe UI", size=10, bold=True, color=TEXT)
    ws.cell(total_row, match_col).border = _total_border()
    ws.cell(total_row, match_col).alignment = Alignment(horizontal="center", vertical="center")
    _apply_number_formats(ws, qb_headers, data_row, total_row, qb_start,
                          {result.qb_mapping["amount"]}, {result.qb_mapping.get("quantity") or ""})
    _apply_number_formats(ws, inf_headers, data_row, total_row, inf_start,
                          {result.inf_mapping["amount"]}, set())
    _set_widths(ws, qb_start, qb_end, header_row, total_row)
    _set_widths(ws, inf_start, inf_end, header_row, total_row)
    ws.column_dimensions[get_column_letter(match_col)].width = 43
    ws.freeze_panes = f"{get_column_letter(inf_start)}{data_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(inf_end)}{final_data_row}"
    ws.print_title_rows = "1:3"
    _prepare_sheet(ws)


def build_unresolved_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Unresolved Exceptions")
    source_headers = list(result.qb_raw.columns)
    headers = source_headers + ["Exception Status", "Reference Amount Difference", "Reviewer Note"]
    candidate_map = result.candidates.set_index("QuickBooks Row ID").to_dict("index") if not result.candidates.empty else {}
    records = []
    for qidx in result.unmatched_qb:
        candidate = candidate_map.get(result.qb_work.at[qidx, QB_ID], {})
        records.append(
            [result.qb_work.at[qidx, col] for col in source_headers]
            + [
                candidate.get("Disposition", "Unmatched QuickBooks"),
                candidate.get("Minimum Amount Difference"),
                "",
            ]
        )
    frame = pd.DataFrame(records, columns=headers)
    end_col = len(headers)
    amount_col_position = source_headers.index(result.qb_mapping["amount"]) + 1
    amounts = [cents_or_zero(result.qb_work.at[idx, AMOUNT_CENTS]) for idx in result.unmatched_qb]
    debits = sum(value for value in amounts if value > 0)
    credits = abs(sum(value for value in amounts if value < 0))
    net = sum(amounts)

    _write_title_band(ws, 1, 1, end_col, "ACCOUNTING EXCEPTIONS | JOURNAL ENTRY SUPPORT", NAVY)
    invalid_unresolved = sum(
        1 for idx in result.unmatched_qb if not valid_cents(result.qb_work.at[idx, AMOUNT_CENTS])
    )
    _write_caption_band(
        ws, 2, 1, end_col,
        f"Net signed support total excludes {invalid_unresolved} row(s) with invalid or missing amounts. "
        f"Review every exception before posting. Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    kpis = [
        ("Unresolved rows", len(result.unmatched_qb), '#,##0'),
        ("Gross debits", cents_to_float(debits), '$#,##0.00;[Red]($#,##0.00);-'),
        ("Credits", cents_to_float(credits), '$#,##0.00;[Red]($#,##0.00);-'),
        ("Proposed JE support total", cents_to_float(net), '$#,##0.00;[Red]($#,##0.00);-'),
    ]
    for idx, (label, value, number_format) in enumerate(kpis):
        start = 1 + idx * 2
        if start > end_col:
            break
        ws.cell(3, start, label)
        ws.cell(4, start, value)
        ws.cell(3, start).font = Font(name="Segoe UI", size=9, bold=True, color=SLATE)
        ws.cell(4, start).font = Font(name="Segoe UI", size=12, bold=True, color=NAVY)
        ws.cell(4, start).number_format = number_format
        for row in (3, 4):
            ws.cell(row, start).fill = PatternFill("solid", fgColor=SLATE_LIGHT)
            ws.cell(row, start).border = _thin_border()

    header_row, data_row = 6, 7
    _write_dataframe_values(ws, frame, header_row, 1)
    _format_header(ws, header_row, 1, end_col, NAVY)
    if len(frame):
        _format_body_block(ws, data_row, data_row + len(frame) - 1, 1, end_col, NAVY_LIGHT)
        duplicate_qb_rows = _duplicate_source_indexes(result, "QuickBooks")
        for offset, qidx in enumerate(result.unmatched_qb):
            row = data_row + offset
            ws.cell(row, source_headers.index(result.qb_mapping["amount"]) + 1).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            ws.cell(row, len(source_headers) + 1).fill = PatternFill("solid", fgColor=AMBER)
            ws.cell(row, len(source_headers) + 1).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row, len(source_headers) + 2).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            if int(qidx) in duplicate_qb_rows:
                _apply_duplicate_style(ws, row, 1, end_col)
    total_row = data_row + len(frame)
    _write_total_row(
        ws, total_row, 1, end_col,
        {result.qb_mapping["amount"]: cents_to_float(net)}, headers,
        "PROPOSED JE SUPPORT TOTAL",
    )
    ws.cell(total_row, amount_col_position).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    ws.column_dimensions[get_column_letter(len(source_headers) + 1)].width = 48
    ws.column_dimensions[get_column_letter(len(source_headers) + 2)].width = 24
    ws.column_dimensions[get_column_letter(len(source_headers) + 3)].width = 36
    _set_widths(ws, 1, len(source_headers), header_row, total_row)
    ws.freeze_panes = f"A{data_row}"
    if len(frame):
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(end_col)}{header_row + len(frame)}"
        note_col = get_column_letter(len(source_headers) + 3)
        validation = DataValidation(
            type="textLength", operator="lessThanOrEqual", formula1="1000", allow_blank=True
        )
        validation.error = "Reviewer notes are limited to 1,000 characters."
        validation.errorTitle = "Note too long"
        ws.add_data_validation(validation)
        validation.add(f"{note_col}{data_row}:{note_col}{header_row + len(frame)}")

    journal_amount = abs(cents_to_float(net))
    je_title_row = total_row + 3
    je_caption_row = je_title_row + 1
    je_header_row = je_title_row + 2
    je_data_row = je_header_row + 1
    je_headers = [
        "Entry Name", "GL Account", "Account Name", "Debit", "Credit", "Entry Basis",
    ]
    je_frame = pd.DataFrame(
        [
            [
                "AC001 Sales Accrual",
                "017-00000-110160.0",
                "Accrued Income",
                journal_amount,
                0.0,
                "Unresolved QuickBooks net exception support",
            ],
            [
                "AC001 Sales Accrual",
                "017-91000-400000-0",
                "Income-Manufacturing",
                0.0,
                journal_amount,
                "Balanced offset",
            ],
        ],
        columns=je_headers,
    )
    _write_title_band(
        ws, je_title_row, 1, end_col,
        "PROPOSED JOURNAL ENTRY | AC001 SALES ACCRUAL",
        SLATE,
    )
    _write_caption_band(
        ws, je_caption_row, 1, end_col,
        "Post only after review and approval. Debit 017-00000-110160.0 Accrued Income and credit "
        "017-91000-400000-0 Income-Manufacturing for the absolute unresolved net amount; evaluate "
        "reversals and negative source values before posting.",
        SLATE,
    )
    _write_dataframe_values(ws, je_frame, je_header_row, 1)
    _format_header(ws, je_header_row, 1, len(je_headers), SLATE)
    _format_body_block(ws, je_data_row, je_data_row + len(je_frame) - 1, 1, len(je_headers), SLATE_LIGHT)
    for row in range(je_data_row, je_data_row + len(je_frame)):
        ws.cell(row, 4).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        ws.cell(row, 5).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    je_total_row = je_data_row + len(je_frame)
    _write_total_row(
        ws, je_total_row, 1, len(je_headers),
        {"Debit": journal_amount, "Credit": journal_amount}, je_headers, "BALANCED TOTAL",
    )
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 24)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 23)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 28)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width or 0, 52)

    duplicate_frame = result.duplicate_analysis
    duplicate_headers = list(duplicate_frame.columns)
    duplicate_title_row = je_total_row + 3
    duplicate_header_row = duplicate_title_row + 2
    duplicate_data_row = duplicate_header_row + 1
    duplicate_end_col = len(duplicate_headers)
    section_end_col = max(end_col, duplicate_end_col)
    _write_title_band(
        ws, duplicate_title_row, 1, section_end_col,
        "DUPLICATE MATCHING-KEY REVIEW | QUICKBOOKS AND INFINIUM", TEAL,
    )
    duplicate_caption = (
        f"{len(duplicate_frame):,} repeated matching-key group(s) identified. Review these source values for "
        "duplicate entries or legitimate repeated transactions."
        if len(duplicate_frame)
        else "No repeated PO/invoice/amount matching keys were identified in either dataset."
    )
    _write_caption_band(ws, duplicate_title_row + 1, 1, section_end_col, duplicate_caption, TEAL)
    _write_dataframe_values(ws, duplicate_frame, duplicate_header_row, 1)
    _format_header(ws, duplicate_header_row, 1, duplicate_end_col, TEAL)
    if len(duplicate_frame):
        duplicate_last_row = duplicate_data_row + len(duplicate_frame) - 1
        _format_body_block(ws, duplicate_data_row, duplicate_last_row, 1, duplicate_end_col, TEAL_LIGHT)
        _apply_number_formats(
            ws, duplicate_headers, duplicate_data_row, duplicate_last_row, 1,
            {"Amount"}, set(),
        )
        for row in range(duplicate_data_row, duplicate_last_row + 1):
            _apply_duplicate_style(ws, row, 1, duplicate_end_col)
        ws.column_dimensions[get_column_letter(duplicate_headers.index("Source Row IDs") + 1)].width = 44
        ws.column_dimensions[get_column_letter(duplicate_headers.index("Reconciliation Status") + 1)].width = 42
    else:
        duplicate_last_row = duplicate_header_row

    fiscal_summary = build_fiscal_exception_summary(result)
    fiscal_headers = list(fiscal_summary.columns)
    fiscal_title_row = duplicate_last_row + 3
    fiscal_caption_row = fiscal_title_row + 1
    fiscal_header_row = fiscal_title_row + 2
    fiscal_data_row = fiscal_header_row + 1
    fiscal_end_col = len(fiscal_headers)
    fiscal_section_end_col = max(end_col, fiscal_end_col)
    selected_period = result.metadata.get("fiscal_period")
    has_fiscal_period = bool(result.qb_mapping.get("period"))
    _write_title_band(
        ws, fiscal_title_row, 1, fiscal_section_end_col,
        (
            "FISCAL-PERIOD EXCEPTION SUMMARY | CURRENT VS PRIOR PERIODS"
            if has_fiscal_period
            else "EXCEPTION SUMMARY | FISCAL PERIOD NOT AVAILABLE"
        ),
        NAVY,
    )
    quantity_note = (
        "Exception quantity is sourced from the mapped QuickBooks quantity column."
        if result.qb_mapping.get("quantity")
        else "No QuickBooks quantity column was mapped; exception quantities are shown as zero."
    )
    _write_caption_band(
        ws, fiscal_caption_row, 1, fiscal_section_end_col,
        (
            f"Selected current reporting period: PD-{int(selected_period):02d}. Every other valid QuickBooks fiscal "
            f"period is classified as a prior-period urgent exception. {quantity_note}"
            if has_fiscal_period and selected_period is not None
            else f"No current reporting period was selected. Exceptions are summarized by source period without current/prior classification. {quantity_note}"
            if has_fiscal_period
            else "No credible QuickBooks fiscal-period identifier was found or mapped. Period-based "
            f"classification is disabled and all exceptions are summarized together. {quantity_note}"
        ),
        NAVY,
    )
    _write_dataframe_values(ws, fiscal_summary, fiscal_header_row, 1)
    _format_header(ws, fiscal_header_row, 1, fiscal_end_col, NAVY)
    if len(fiscal_summary):
        fiscal_last_row = fiscal_data_row + len(fiscal_summary) - 1
        _format_body_block(ws, fiscal_data_row, fiscal_last_row, 1, fiscal_end_col, NAVY_LIGHT)
        for offset, classification in enumerate(fiscal_summary["Period Classification"], start=fiscal_data_row):
            fill = (
                RED_LIGHT
                if classification == "Prior-Period Urgent Exception"
                else GREEN_LIGHT
                if classification == "Current Reporting Period"
                else AMBER
            )
            for col in range(1, fiscal_end_col + 1):
                ws.cell(offset, col).fill = PatternFill("solid", fgColor=fill)
        _apply_number_formats(
            ws, fiscal_headers, fiscal_data_row, fiscal_last_row, 1,
            {"Net Exception Amount"}, {"Exception Count", "Exception Quantity"},
        )
    else:
        fiscal_last_row = fiscal_header_row
    fiscal_total_row = fiscal_last_row + 1
    _write_total_row(
        ws, fiscal_total_row, 1, fiscal_end_col,
        {
            "Exception Count": float(fiscal_summary["Exception Count"].sum()) if len(fiscal_summary) else 0,
            "Exception Quantity": float(fiscal_summary["Exception Quantity"].sum()) if len(fiscal_summary) else 0,
            "Net Exception Amount": float(fiscal_summary["Net Exception Amount"].sum()) if len(fiscal_summary) else 0,
        },
        fiscal_headers,
        "TOTAL EXCEPTIONS",
    )
    _set_widths(ws, 1, fiscal_end_col, fiscal_header_row, fiscal_total_row)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 34)
    ws.print_title_rows = "1:6"
    _prepare_sheet(ws)


def build_product_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Product Aggregate Summary")
    frame = result.product_summary
    headers = ["Product Name", "Product Quantity", "Product Value"]
    _write_title_band(ws, 1, 1, 3, "PRODUCT AGGREGATE SUMMARY", NAVY)
    selected_period = result.metadata.get("fiscal_period")
    period_scope = (
        f"Only primary QuickBooks rows from Period {int(selected_period):02d} are included."
        if selected_period is not None
        else "All primary QuickBooks fiscal periods are included."
    )
    _write_caption_band(
        ws, 2, 1, 3,
        "Products are classified from the QuickBooks description column. "
        f"{period_scope} This classification does not influence data matching. "
        f"Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    _write_dataframe_values(ws, frame.reindex(columns=headers), 3, 1)
    _format_header(ws, 3, 1, 3, NAVY)
    if len(frame):
        _format_body_block(ws, 4, 3 + len(frame), 1, 3, NAVY_LIGHT)
    total_row = 4 + len(frame)
    totals = {
        "Product Quantity": float(frame["Product Quantity"].sum()) if not frame.empty else 0,
        "Product Value": float(frame["Product Value"].sum()) if not frame.empty else 0,
    }
    _write_total_row(ws, total_row, 1, 3, totals, headers)
    _apply_number_formats(ws, headers, 4, total_row, 1, {"Product Value"}, {"Product Quantity"})
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A4"
    if len(frame):
        ws.auto_filter.ref = f"A3:C{3 + len(frame)}"
    _prepare_sheet(ws, landscape=False)


def _save_workbook_bytes(wb: Workbook) -> bytes:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    _autofit_workbook_columns(wb)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _apply_workbook_run_metadata(wb: Workbook, result: ReconciliationResult) -> None:
    central_timestamp = result.run_timestamp.astimezone(CENTRAL_TIMEZONE)
    excel_timestamp = central_timestamp.replace(tzinfo=None)
    display_timestamp = format_central_timestamp(central_timestamp)
    # Visible report captions carry the controlled Central timestamp. Core
    # properties use the same run time, while page headers/footers are blank so
    # Excel cannot surface a stale or locale-generated tag on printed sheets.
    wb.properties.created = excel_timestamp
    wb.properties.modified = excel_timestamp
    existing_description = wb.properties.description or ""
    wb.properties.description = (
        f"{existing_description} Generated {display_timestamp}. Run ID: {result.run_id}."
    ).strip()
    for ws in wb.worksheets:
        for section in (
            ws.oddHeader, ws.evenHeader, ws.firstHeader,
            ws.oddFooter, ws.evenFooter, ws.firstFooter,
        ):
            section.left.text = None
            section.center.text = None
            section.right.text = None


def build_primary_workbook(result: ReconciliationResult) -> bytes:
    wb = Workbook()
    wb.properties.creator = "Sales Reconciliation Application"
    wb.properties.title = f"Sales Reconciliation {result.run_id}"
    wb.properties.subject = "QuickBooks to Infinium reconciliation and journal-entry support"
    wb.properties.description = "Four-sheet accounting workpaper generated from one controlled reconciliation run."
    build_raw_data_sheet(wb, result)
    build_reconciled_data_sheet(wb, result)
    build_unresolved_sheet(wb, result)
    build_product_sheet(wb, result)
    _apply_workbook_run_metadata(wb, result)
    return _save_workbook_bytes(wb)


def detailed_ledger_dataframe(result: ReconciliationResult) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    qb_display, inf_display, _ = _paired_display_frames(result)
    for position, stored_record in enumerate(result.paired_rows):
        record = _resolved_paired_record(result, stored_record)
        output = {
            "Run ID": result.run_id,
            "Section": record["Section"],
            "Match ID": record["Match ID"],
            "Match Result": record["Match Result"],
            "Confidence": record["Confidence"],
            "Group Sequence": record["Group Sequence"],
            "Assessment Explanation": record["Explanation"],
        }
        qidx, iidx = record["QB Index"], record["Infinium Index"]
        qb_frame = (
            result.qb_secondary_work
            if record.get("QB Record Scope") == "Historical" else result.qb_work
        )
        inf_frame = (
            result.inf_secondary_work
            if record.get("Infinium Record Scope") == "Historical" else result.inf_work
        )
        output["QB | Source Row ID"] = (
            qb_frame.at[qidx, QB_ID] if qidx is not None and qb_frame is not None else None
        )
        for header, value in qb_display.iloc[position].items():
            output[f"QB | {header}"] = value
        output["INF | Source Row ID"] = (
            inf_frame.at[iidx, INF_ID] if iidx is not None and inf_frame is not None else None
        )
        for header, value in inf_display.iloc[position].items():
            output[f"INF | {header}"] = value
        records.append(output)
    return pd.DataFrame(records)


def _add_standard_data_sheet(
    wb: Workbook,
    name: str,
    title: str,
    caption: str,
    frame: pd.DataFrame,
    header_color: str = NAVY,
    chart_column: Optional[str] = None,
) -> Any:
    ws = wb.create_sheet(name)
    end_col = max(len(frame.columns), 1)
    _write_title_band(ws, 1, 1, end_col, title, header_color)
    _write_caption_band(ws, 2, 1, end_col, caption, header_color)
    if frame.empty and len(frame.columns) == 0:
        frame = pd.DataFrame({"Result": ["No records"]})
        end_col = 1
    _write_dataframe_values(ws, frame, 3, 1)
    _format_header(ws, 3, 1, len(frame.columns), header_color)
    if len(frame):
        _format_body_block(ws, 4, 3 + len(frame), 1, len(frame.columns), NAVY_LIGHT)
        _apply_number_formats(ws, list(frame.columns), 4, 3 + len(frame), 1, set(), set())
    _set_widths(ws, 1, len(frame.columns), 3, max(3 + len(frame), 3), maximum=48)
    wrap_terms = ("EXPLANATION", "REQUIREMENT", "DISPOSITION", "MATCH RESULT", "CANDIDATE IDS")
    for col, header in enumerate(frame.columns, 1):
        if any(term in str(header).upper() for term in wrap_terms):
            ws.column_dimensions[get_column_letter(col)].width = 42
            for row in range(4, 4 + len(frame)):
                ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 20, 32)
    ws.freeze_panes = "A4"
    if len(frame):
        ws.auto_filter.ref = f"A3:{get_column_letter(len(frame.columns))}{3 + len(frame)}"
    if chart_column and chart_column in frame.columns and len(frame):
        data_col = list(frame.columns).index(chart_column) + 1
        category_col = 1
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = f"{chart_column} by Match Method"
        chart.y_axis.title = "Match Method"
        chart.x_axis.title = chart_column
        chart.height = 7.5
        chart.width = 15
        data = Reference(ws, min_col=data_col, min_row=3, max_row=3 + len(frame))
        categories = Reference(ws, min_col=category_col, min_row=4, max_row=3 + len(frame))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        ws.add_chart(chart, f"{get_column_letter(len(frame.columns) + 2)}3")
    _prepare_sheet(ws)
    return ws


def build_analytics_summary_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    _write_title_band(ws, 1, 1, 8, "RECONCILIATION ANALYTICS | EXECUTIVE SUMMARY", NAVY)
    _write_caption_band(
        ws, 2, 1, 8,
        f"Run ID: {result.run_id} | Generated {format_central_timestamp(result.run_timestamp)} | Control status: {result.metrics['Control Status']}",
        NAVY,
    )
    cards = [
        ("QuickBooks rows", result.metrics["QuickBooks Rows"], '#,##0'),
        ("Infinium rows", result.metrics["Infinium Rows"], '#,##0'),
        ("QB match rate", result.metrics["QuickBooks Match Rate by Row"], '0.0%'),
        ("Unresolved QB amount", result.metrics["Unresolved QuickBooks Amount"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("QB source total", result.metrics["QuickBooks Source Total"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Infinium source total", result.metrics["Infinium Source Total"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Matched amount difference", result.metrics["Matched Amount Difference"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Control status", result.metrics["Control Status"], '@'),
    ]
    for idx, (label, value, number_format) in enumerate(cards):
        row = 4 if idx < 4 else 7
        col = 1 + (idx % 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col, label)
        ws.cell(row + 1, col, excel_safe(value))
        for r in (row, row + 1):
            for c in (col, col + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=SLATE_LIGHT)
                ws.cell(r, c).border = _thin_border()
        ws.cell(row, col).font = Font(name="Segoe UI", size=9, bold=True, color=SLATE)
        ws.cell(row + 1, col).font = Font(
            name="Segoe UI", size=13, bold=True,
            color=NAVY if value != "FAIL" else "9C0006",
        )
        ws.cell(row + 1, col).number_format = number_format

    start = 11
    ws.cell(start, 1, "MODEL STATUS")
    ws.cell(start, 2, result.metrics["Control Status"])
    ws.cell(start, 1).font = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=SLATE)
    ws.cell(start, 2).font = Font(name="Segoe UI", size=11, bold=True, color=TEXT)
    ws.cell(start, 2).fill = PatternFill("solid", fgColor=GREEN_LIGHT if result.metrics["Control Status"] == "PASS" else RED_LIGHT)
    _write_dataframe_values(ws, result.controls, start + 2, 1)
    _format_header(ws, start + 2, 1, len(result.controls.columns), SLATE)
    _format_body_block(ws, start + 3, start + 2 + len(result.controls), 1, len(result.controls.columns), SLATE_LIGHT)
    _apply_number_formats(ws, list(result.controls.columns), start + 3, start + 2 + len(result.controls), 1, set(), set())
    status_col = list(result.controls.columns).index("Status") + 1
    if len(result.controls):
        green = PatternFill("solid", fgColor=GREEN_LIGHT)
        red = PatternFill("solid", fgColor=RED_LIGHT)
        ws.conditional_formatting.add(
            f"{get_column_letter(status_col)}{start + 3}:{get_column_letter(status_col)}{start + 2 + len(result.controls)}",
            FormulaRule(formula=[f'{get_column_letter(status_col)}{start + 3}="PASS"'], fill=green),
        )
        ws.conditional_formatting.add(
            f"{get_column_letter(status_col)}{start + 3}:{get_column_letter(status_col)}{start + 2 + len(result.controls)}",
            FormulaRule(formula=[f'{get_column_letter(status_col)}{start + 3}="FAIL"'], fill=red),
        )
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A3"
    _prepare_sheet(ws)


def build_analytics_workbook(result: ReconciliationResult) -> bytes:
    wb = Workbook()
    wb.properties.creator = "Sales Reconciliation Application"
    wb.properties.title = f"Sales Reconciliation Analytics {result.run_id}"
    wb.properties.subject = "Detailed matching evidence and reconciliation controls"
    build_analytics_summary_sheet(wb, result)
    _add_standard_data_sheet(
        wb, "Match Method Summary", "MATCH METHOD SUMMARY",
        "Distribution of matched and unresolved rows. Percentages use QuickBooks row count as the denominator.",
        result.method_summary, NAVY, chart_column="QuickBooks Rows",
    )
    _add_standard_data_sheet(
        wb, "Detailed Match Ledger", "DETAILED MATCH Ledger",
        "One record per displayed reconciliation line. Historical clearances show the specific accepted prior-period row on the opposing side; unused historical rows are excluded.",
        detailed_ledger_dataframe(result), SLATE,
    )
    _add_standard_data_sheet(
        wb, "Normalization Detail", "NORMALIZATION DETAIL",
        "Original source values and the exact normalized references and signed amounts considered by the matching engine.",
        result.normalization, TEAL,
    )
    _add_standard_data_sheet(
        wb, "Match Assessment", "MATCH ASSESSMENT",
        "One row per accepted match group or unresolved QuickBooks decision, including criteria and evidence.",
        result.assessments, NAVY,
    )
    _add_standard_data_sheet(
        wb, "Historical Clearances", "SECONDARY HISTORICAL CLEARANCES",
        "Accepted strict matches between an unresolved primary row and the opposing historical source. Unused secondary rows are intentionally excluded.",
        result.historical_clearances, TEAL,
    )
    _add_standard_data_sheet(
        wb, "Exception Analysis", "EXCEPTION ANALYSIS",
        "Unresolved population summarized by source period, reason, and source ledger. Period is reporting metadata only.",
        result.exception_analysis, NAVY,
    )
    _add_standard_data_sheet(
        wb, "Rules and Run Config", "RULES AND RUN CONFIGURATION",
        "Exact rules, file fingerprints, mappings, versions, and parameters used for this run.",
        pd.concat(
            [
                result.config,
                pd.DataFrame([{"Setting": "", "Value": ""}, {"Setting": "MATCHING RULES", "Value": ""}]),
                result.rules.rename(columns={
                    "Priority": "Setting",
                    "Rule": "Value",
                    "Automatic": "Automatic",
                    "Requirement": "Requirement",
                }),
            ],
            ignore_index=True,
        ),
        SLATE,
    )
    _apply_workbook_run_metadata(wb, result)
    return _save_workbook_bytes(wb)


# ---------------------------------------------------------------------------
# Streamlit interface
# ---------------------------------------------------------------------------


def load_app_css() -> None:
    """Load the stylesheet located beside this application module."""
    css_path = Path(__file__).resolve().with_name("style.css")
    try:
        css = css_path.read_text(encoding="utf-8")
        css = (
            css.replace("\u00a0", " ")
            .replace("\u2007", " ")
            .replace("\u202f", " ")
        )
    except OSError as exc:
        st.warning(f"The application stylesheet could not be loaded: {exc}")
        return

    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)


def format_currency(value: Any) -> str:
    numeric = float(value or 0)
    return f"${numeric:,.2f}" if numeric >= 0 else f"(${abs(numeric):,.2f})"


def render_kpi(label: str, value: str, subtitle: str = "") -> None:
    st.markdown(
        f"""
        <div class="rec-card">
            <div class="rec-kpi-label">{label}</div>
            <div class="rec-kpi-value">{value}</div>
            <div class="rec-kpi-sub">{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def display_primary_preview(result: ReconciliationResult) -> pd.DataFrame:
    qb_display, inf_display, methods = _paired_display_frames(result)
    qb_display = qb_display.add_prefix("QB | ")
    inf_display = inf_display.add_prefix("INF | ")
    return pd.concat(
        [qb_display, pd.DataFrame({"Match Result": methods}), inf_display], axis=1
    )


def render_limited_dataframe(
    frame: pd.DataFrame,
    *,
    height: Optional[int] = None,
    row_limit: int = 2000,
) -> None:
    """Keep browser previews responsive while preserving full Excel outputs."""
    displayed = frame.head(row_limit)
    kwargs: dict[str, Any] = {
        "use_container_width": True,
        "hide_index": True,
    }
    if height is not None:
        kwargs["height"] = height
    st.dataframe(displayed, **kwargs)
    if len(frame) > row_limit:
        st.caption(
            f"Showing the first {row_limit:,} of {len(frame):,} rows. "
            "The complete population is retained in the downloadable workbook."
        )


def show_toast_once(state_key: str, message: str, icon: str = "✅") -> None:
    """Show a native toast once without repeating it on every app rerun."""
    if not st.session_state.get(state_key, False):
        st.toast(message, icon=icon)
        st.session_state[state_key] = True


def attention_tab_label(label: str, count: int) -> str:
    """Add a compatible text badge only when a tab has review items."""
    return f"{label} ({count:,})" if count > 0 else label


def run_id_for_inputs(*source_hashes: str) -> tuple[str, datetime]:
    now_utc = datetime.now(timezone.utc)
    now_central = now_utc.astimezone(CENTRAL_TIMEZONE)
    digest_source = "|".join([*source_hashes, now_utc.isoformat()])
    digest = hashlib.sha256(digest_source.encode()).hexdigest()[:8].upper()
    return f"REC-{now_central.strftime('%Y%m%d-%H%M%S')}-{digest}", now_central


def clear_results_if_signature_changed(signature: tuple[Any, ...]) -> None:
    if st.session_state.get("input_signature") != signature:
        st.session_state.input_signature = signature
        st.session_state.pop("reconciliation_result", None)
        st.session_state.pop("primary_workbook", None)
        st.session_state.pop("analytics_workbook", None)


def render_source_status(filename: Optional[str], next_step: str) -> None:
    """Renders the file ingestion status directly beneath the file uploader in the main interface."""
    if filename:
        st.markdown(
            f'<div class="source-status complete">Loaded: {html.escape(filename)}</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="source-status active">Current step: {next_step}</div>',
            unsafe_allow_html=True,
        )


def render_upload_source_heading(
    step: int,
    label: str,
    logo_uri: str,
    logo_class: str,
) -> None:
    """Renders the file upload headings in the main interface columns."""
    st.markdown(
        f'<div class="upload-source-heading">'
        f'<div class="upload-source-copy"><span class="upload-step-number">{step}</span>'
        f'<span class="upload-source-title">{html.escape(label)}</span></div>'
        f'<img class="upload-source-logo {html.escape(logo_class)}" '
        f'src="{logo_uri}" alt="{html.escape(label)} logo">'
        f'</div>',
        unsafe_allow_html=True,
    )


def render_ingestion_flow(
    qb_loaded: bool,
    inf_loaded: bool,
    mappings_ready: bool,
    reconciliation_complete: bool,
    source_validation_failed: bool = False,
) -> None:
    step1_class = "complete" if qb_loaded else "active"
    step1_status = "Upload confirmed" if qb_loaded else "Upload QuickBooks first"
    if inf_loaded:
        step2_class, step2_status = "complete", "Upload confirmed"
    elif qb_loaded:
        step2_class, step2_status = "active", "Upload Infinium next"
    else:
        step2_class, step2_status = "upcoming", "Available after QuickBooks"
    if reconciliation_complete:
        step3_class, step3_status = "complete", "Controls validated"
    elif qb_loaded and inf_loaded and source_validation_failed:
        step3_class, step3_status = "active", "Resolve source validation"
    elif qb_loaded and inf_loaded and mappings_ready:
        step3_class, step3_status = "active", "Ready to run"
    elif qb_loaded and inf_loaded:
        step3_class, step3_status = "active", "Confirm required mappings"
    else:
        step3_class, step3_status = "upcoming", "Waiting for both files"
    if reconciliation_complete:
        step4_class, step4_status = "active", "Review and download"
    else:
        step4_class, step4_status = "upcoming", "Available after reconciliation"

    steps = [
        (1, "Ingest QuickBooks", step1_status, step1_class),
        (2, "Ingest Infinium", step2_status, step2_class),
        (3, "Run Reconciliation", step3_status, step3_class),
        (4, "Review Results", step4_status, step4_class),
    ]
    # Keep this markup on one physical line. Markdown treats indented HTML
    # after a blank line as a code block, which previously rendered card 1 but
    # displayed cards 2-4 as literal <div> text.
    cards = "".join(
        f'<div class="ingestion-step {css_class}">'
        f'<span class="ingestion-number">{number}</span>'
        f'<span class="ingestion-title">{title}</span>'
        f'<div class="ingestion-status">{status}</div>'
        f'</div>'
        for number, title, status, css_class in steps
    )
    flow_html = f'<div class="ingestion-flow">{cards}</div>'
    st.markdown(flow_html, unsafe_allow_html=True)


def render_pre_execution_controls(
    qb_raw: pd.DataFrame,
    inf_raw: pd.DataFrame,
    qb_mapping: dict[str, Optional[str]],
    inf_mapping: dict[str, Optional[str]],
) -> None:
    qb_total = numeric_sum(qb_raw[qb_mapping["amount"]])
    inf_total = numeric_sum(inf_raw[inf_mapping["amount"]])
    st.subheader("Source verification")
    cols = st.columns(4)
    with cols[0]:
        render_kpi("QuickBooks rows", f"{len(qb_raw):,}", format_currency(qb_total))
    with cols[1]:
        render_kpi("Infinium rows", f"{len(inf_raw):,}", format_currency(inf_total))
    with cols[2]:
        render_kpi("Row difference", f"{len(qb_raw) - len(inf_raw):,}", "QB minus Infinium")
    with cols[3]:
        render_kpi("Source value difference", format_currency(qb_total - inf_total), "QB minus Infinium")
    st.caption(
        "Pre-reconciliation control totals include only QuickBooks rows populated in every field other than Quantity and Amount. "
        "Retained rows with invalid amounts are flagged and remain unreconciled."
    )


def render_result(result: ReconciliationResult) -> None:
    metrics = result.metrics
    control_status = metrics["Control Status"]
    if control_status == "PASS":
        show_toast_once(
            f"reconciliation_complete_{result.run_id}",
            f"Reconciliation complete. All controls passed. Run ID: {result.run_id}",
        )
    else:
        st.error("Reconciliation completed with failed controls. Downloads are withheld until controls pass.")

    failed_controls = int(result.controls["Status"].ne("PASS").sum())
    invalid_amount_rows = int(
        metrics["Invalid QuickBooks Amounts"] + metrics["Invalid Infinium Amounts"]
    )
    unmatched_qb = int(metrics["Unresolved QuickBooks Rows"])
    unmatched_inf = int(metrics["Unmatched Infinium Rows"])
    unresolved_duplicate_groups = 0
    if not result.duplicate_analysis.empty:
        unresolved_duplicate_groups = int(
            result.duplicate_analysis["Reconciliation Status"]
            .ne("All rows matched through a more specific unique key")
            .sum()
        )

    tab_labels = [
        attention_tab_label("Overview", failed_controls + invalid_amount_rows),
        attention_tab_label("Reconciled View", unmatched_qb + unmatched_inf),
        attention_tab_label("Exception Review", unmatched_qb),
        attention_tab_label("Analytics", unresolved_duplicate_groups),
        "Downloads",
    ]
    overview_tab, reconciled_tab, exceptions_tab, analytics_tab, downloads_tab = st.tabs(
        tab_labels
    )
    with overview_tab:
        cols = st.columns(4)
        with cols[0]:
            render_kpi(
                "QB match rate",
                f"{metrics['QuickBooks Match Rate by Row']:.1%}",
                f"{metrics['Matched QuickBooks Rows']:,} of {metrics['QuickBooks Rows']:,} rows",
            )
        with cols[1]:
            render_kpi(
                "Unresolved QB",
                format_currency(metrics["Unresolved QuickBooks Amount"]),
                f"{metrics['Unresolved QuickBooks Rows']:,} transactions",
            )
        with cols[2]:
            render_kpi(
                "Matched control difference",
                format_currency(metrics["Matched Amount Difference"]),
                "QuickBooks minus Infinium",
            )
        with cols[3]:
            render_kpi(
                "Control status",
                control_status,
                f"{len(result.controls)} required controls",
            )
        st.markdown("#### Required controls")
        st.dataframe(result.controls, use_container_width=True, hide_index=True)
        if metrics.get("Historical Clearances", 0):
            st.info(
                f"Historical secondary data cleared {metrics['Historical Clearances']:,} opposing-primary "
                "exception item(s). "
                f"{metrics['QuickBooks Secondary Rows Ignored'] + metrics['Infinium Secondary Rows Ignored']:,} "
                "unused secondary row(s) were excluded from exception reporting."
            )
        if metrics["Invalid QuickBooks Amounts"] or metrics["Invalid Infinium Amounts"]:
            st.markdown(
                '<div class="data-quality-banner">Data quality review: '
                f"{metrics['Invalid QuickBooks Amounts']} QuickBooks and "
                f"{metrics['Invalid Infinium Amounts']} Infinium rows have invalid or missing amounts. "
                "Those rows were not automatically matched.</div>",
                unsafe_allow_html=True,
            )

    with reconciled_tab:
        st.caption(
            "Matched rows appear first. Unmatched QuickBooks and unmatched Infinium rows are kept in separate "
            "sections so unrelated exceptions are never displayed as a pair. Historical clearances display the "
            "specific accepted prior-period row on the opposing side and label it in Record Context. Unused "
            "historical rows are never added to this ledger."
        )
        render_limited_dataframe(display_primary_preview(result), height=520)

    with exceptions_tab:
        unresolved = result.qb_work.loc[result.unmatched_qb, list(result.qb_raw.columns)].copy()
        if not result.candidates.empty:
            candidate_map = result.candidates.set_index("QuickBooks Row ID").to_dict("index")
            unresolved["Exception Status"] = [
                candidate_map.get(result.qb_work.at[idx, QB_ID], {}).get(
                    "Disposition", "Unmatched QuickBooks"
                )
                for idx in result.unmatched_qb
            ]
            unresolved["Reference Amount Difference"] = [
                candidate_map.get(result.qb_work.at[idx, QB_ID], {}).get(
                    "Minimum Amount Difference"
                )
                for idx in result.unmatched_qb
            ]
        if unresolved.empty:
            st.caption("No unresolved QuickBooks exceptions remain.")
        else:
            render_limited_dataframe(unresolved, height=440)
        st.metric("Proposed journal-entry support total", format_currency(metrics["Unresolved QuickBooks Amount"]))
        st.caption("This is the net signed amount of unresolved QuickBooks transactions. Review credits and reversals before posting.")

    with analytics_tab:
        left, right = st.columns(2)
        with left:
            st.markdown("#### Match-method distribution")
            st.dataframe(result.method_summary, use_container_width=True, hide_index=True)
        with right:
            st.markdown("#### Exception analysis")
            st.dataframe(result.exception_analysis, use_container_width=True, hide_index=True)
        with st.expander("Normalization and assessment detail", expanded=False):
            st.dataframe(result.assessments, use_container_width=True, hide_index=True, height=320)
    with downloads_tab:
        st.markdown("#### Accounting workpaper")
        st.caption("Four sheets: Raw Data, Reconciled Data, Unresolved Exceptions, and Product Aggregate Summary.")
        if "primary_workbook" not in st.session_state:
            if st.button(
                "Prepare Sales Reconciliation",
                type="primary",
                use_container_width=True,
                key=f"prepare_primary_{result.run_id}",
            ):
                try:
                    with st.spinner("Preparing the accounting workpaper..."):
                        st.session_state.primary_workbook = build_primary_workbook(result)
                    st.toast("Accounting workpaper prepared.", icon="✅")
                except Exception as exc:
                    st.error(f"The accounting workpaper could not be prepared: {exc}")
        if "primary_workbook" in st.session_state:
            st.download_button(
                "Download Sales Reconciliation",
                data=st.session_state.primary_workbook,
                file_name=f"Sales_Reconciliation_{result.run_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
        st.markdown("#### Detailed analytics")
        st.caption("Optional evidence package with normalization, assessments, controls, and exact run configuration.")
        if "analytics_workbook" not in st.session_state:
            if st.button(
                "Prepare Reconciliation Analytics",
                use_container_width=True,
                key=f"prepare_analytics_{result.run_id}",
            ):
                try:
                    with st.spinner("Preparing the detailed analytics workbook..."):
                        st.session_state.analytics_workbook = build_analytics_workbook(result)
                    st.toast("Analytics workbook prepared.", icon="✅")
                except Exception as exc:
                    st.error(f"The analytics workbook could not be prepared: {exc}")
        if "analytics_workbook" in st.session_state:
            st.download_button(
                "Download Reconciliation Analytics",
                data=st.session_state.analytics_workbook,
                file_name=f"Sales_Reconciliation_Analytics_{result.run_id}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


def main() -> None:
    st.set_page_config(page_title="Sales Reconciliation", layout="wide")
    load_app_css()
    st.markdown(
        """
        <div class="rec-title">
            <h1>Sales Reconciliation</h1>
            <p>QuickBooks-to-Infinium matching, data summarization, and reconciliation analytics</p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # Empty container placed at the top so we can dynamically inject the final ingestion flow state
    progress_container = st.container()

    st.markdown("### Upload Data Sources")
    upload_col1, upload_col2 = st.columns(2)

    with upload_col1:
        render_upload_source_heading(1, "Import QuickBooks file", QUICKBOOKS_LOGO_URI, "quickbooks")
        qb_file = st.file_uploader(
            "Primary QuickBooks import",
            type=["xlsx", "csv"],
            key="qb_file",
            help="Current QuickBooks sales import used as a primary reconciliation source.",
        )
        render_source_status(qb_file.name if qb_file else None, "Upload the QuickBooks import")

    with upload_col2:
        render_upload_source_heading(2, "Import Infinium file", INFOR_LOGO_URI, "infor")
        inf_file = st.file_uploader(
            "Primary Infinium import",
            type=["xlsx", "csv"],
            key="inf_file",
            help="Current Infinium sales import used as a primary reconciliation source.",
        )
        if inf_file:
            render_source_status(inf_file.name, "")
        elif qb_file:
            render_source_status(None, "Upload the Infinium import")
        else:
            st.markdown(
                '<div class="source-status pending">Upcoming: Infinium upload</div>',
                unsafe_allow_html=True,
            )

    qb_secondary_file = None
    inf_secondary_file = None

    if qb_file and inf_file:
        st.markdown("<br>### Optional Historical / Secondary Data", unsafe_allow_html=True)
        st.caption(
            "Optional historical files are background matching data only. They may "
            "clear timing differences in the opposing primary file, but unused rows "
            "will never generate exceptions or appear as reconciliation items."
        )
        sec_col1, sec_col2 = st.columns(2)
        
        with sec_col1:
            qb_secondary_file = st.file_uploader(
                "Historical QuickBooks data",
                type=["xlsx", "csv"],
                key="qb_secondary_file",
                help=(
                    "Used only to match and clear unresolved items from the primary "
                    "Infinium upload. Unused QuickBooks historical rows are ignored."
                ),
            )
            if qb_secondary_file:
                render_source_status(qb_secondary_file.name, "")

        with sec_col2:
            inf_secondary_file = st.file_uploader(
                "Historical Infinium data",
                type=["xlsx", "csv"],
                key="inf_secondary_file",
                help=(
                    "Used only to match and clear unresolved items from the primary "
                    "QuickBooks upload. Unused Infinium historical rows are ignored."
                ),
            )
            if inf_secondary_file:
                render_source_status(inf_secondary_file.name, "")

    st.markdown("---")

    # Shift configuration cleanly into the sidebar
    st.sidebar.markdown("## Configuration Settings")

    if not qb_file or not inf_file:
        for state_key in ("reconciliation_result", "primary_workbook", "analytics_workbook"):
            st.session_state.pop(state_key, None)
            
        with progress_container:
            render_ingestion_flow(
                qb_loaded=qb_file is not None,
                inf_loaded=inf_file is not None,
                mappings_ready=False,
                reconciliation_complete=False,
            )
            
        next_source = "Infinium" if qb_file else "QuickBooks"
        st.info(
            f"**Action Required:** Upload Panhandle Pure {next_source} sales export to continue. "
            "Files are processed locally in the running application session."
        )
        
        st.success(
            "**Automatic matching sequence**\n\n"
            "1. **Unique PO + invoice + exact signed amount**\n"
            "2. **Unique PO + exact signed amount**\n"
            "3. **Unique invoice + exact signed amount**\n\n"
            "---\n"
            "Every accepted match is one-to-one. Grouped relationships, amount differences of any size, "
            "multiple candidates, and invalid amounts remain unresolved for review."
        )
        return

    qb_bytes, inf_bytes = qb_file.getvalue(), inf_file.getvalue()
    qb_secondary_bytes = qb_secondary_file.getvalue() if qb_secondary_file else None
    inf_secondary_bytes = inf_secondary_file.getvalue() if inf_secondary_file else None
    qb_hash, inf_hash = file_sha256(qb_bytes), file_sha256(inf_bytes)
    qb_secondary_hash = file_sha256(qb_secondary_bytes) if qb_secondary_bytes else "NONE"
    inf_secondary_hash = file_sha256(inf_secondary_bytes) if inf_secondary_bytes else "NONE"
    try:
        qb_sheets = list_source_sheets(qb_bytes, qb_file.name)
        inf_sheets = list_source_sheets(inf_bytes, inf_file.name)
        qb_secondary_sheets = (
            list_source_sheets(qb_secondary_bytes, qb_secondary_file.name)
            if qb_secondary_bytes and qb_secondary_file else ()
        )
        inf_secondary_sheets = (
            list_source_sheets(inf_secondary_bytes, inf_secondary_file.name)
            if inf_secondary_bytes and inf_secondary_file else ()
        )
    except Exception as exc:
        st.error(f"The application could not inspect the uploaded workbook structure: {exc}")
        return

    qb_sheet_name: Optional[str] = qb_sheets[0] if qb_sheets else None
    inf_sheet_name: Optional[str] = inf_sheets[0] if inf_sheets else None
    qb_secondary_sheet_name: Optional[str] = (
        qb_secondary_sheets[0] if qb_secondary_sheets else None
    )
    inf_secondary_sheet_name: Optional[str] = (
        inf_secondary_sheets[0] if inf_secondary_sheets else None
    )
    if any(len(sheets) > 1 for sheets in (
        qb_sheets, inf_sheets, qb_secondary_sheets, inf_secondary_sheets
    )):
        st.sidebar.markdown("### Worksheet selection")
    if len(qb_sheets) > 1:
        st.sidebar.warning(
            f"The QuickBooks workbook contains {len(qb_sheets)} worksheets. Select the worksheet containing the reconciliation data."
        )
        qb_sheet_name = st.sidebar.selectbox(
            "QuickBooks worksheet",
            options=list(qb_sheets),
            key=f"qb_sheet_{qb_hash[:12]}",
        )
    if len(inf_sheets) > 1:
        st.sidebar.warning(
            f"The Infinium workbook contains {len(inf_sheets)} worksheets. Select the worksheet containing the reconciliation data."
        )
        inf_sheet_name = st.sidebar.selectbox(
            "Infinium worksheet",
            options=list(inf_sheets),
            key=f"inf_sheet_{inf_hash[:12]}",
        )
    if len(qb_secondary_sheets) > 1:
        qb_secondary_sheet_name = st.sidebar.selectbox(
            "QuickBooks secondary worksheet",
            options=list(qb_secondary_sheets),
            key=f"qb_secondary_sheet_{qb_secondary_hash[:12]}",
        )
    if len(inf_secondary_sheets) > 1:
        inf_secondary_sheet_name = st.sidebar.selectbox(
            "Infinium secondary worksheet",
            options=list(inf_secondary_sheets),
            key=f"inf_secondary_sheet_{inf_secondary_hash[:12]}",
        )

    qb_sheet_key = hashlib.sha256((qb_sheet_name or "CSV").encode()).hexdigest()[:8]
    inf_sheet_key = hashlib.sha256((inf_sheet_name or "CSV").encode()).hexdigest()[:8]
    qb_secondary_sheet_key = hashlib.sha256(
        (qb_secondary_sheet_name or "CSV").encode()
    ).hexdigest()[:8]
    inf_secondary_sheet_key = hashlib.sha256(
        (inf_secondary_sheet_name or "CSV").encode()
    ).hexdigest()[:8]
    key_suffix = (
        f"{qb_hash[:8]}_{qb_sheet_key}_{inf_hash[:8]}_{inf_sheet_key}_"
        f"{qb_secondary_hash[:8]}_{qb_secondary_sheet_key}_"
        f"{inf_secondary_hash[:8]}_{inf_secondary_sheet_key}"
    )
    try:
        detected_qb_header = detect_header_row(
            qb_bytes, qb_file.name, "QB", qb_sheet_name
        )
        detected_inf_header = detect_header_row(
            inf_bytes, inf_file.name, "INF", inf_sheet_name
        )
        detected_qb_secondary_header = (
            detect_header_row(
                qb_secondary_bytes, qb_secondary_file.name, "QB",
                qb_secondary_sheet_name,
            )
            if qb_secondary_bytes and qb_secondary_file else None
        )
        detected_inf_secondary_header = (
            detect_header_row(
                inf_secondary_bytes, inf_secondary_file.name, "INF",
                inf_secondary_sheet_name,
            )
            if inf_secondary_bytes and inf_secondary_file else None
        )
    except Exception as exc:
        st.error(f"The application could not inspect the selected source worksheets: {exc}")
        return

    st.sidebar.markdown("### Import settings")
    with st.sidebar.expander("Header rows", expanded=False):
        st.caption("Excel row numbers are one-based. The detected rows can be overridden if needed.")
        qb_header_number = int(
            st.number_input("QuickBooks header row", min_value=1, max_value=100,
                            value=detected_qb_header + 1, step=1,
                            key=f"qb_header_{qb_hash[:8]}_{qb_sheet_key}")
        )
        inf_header_number = int(
            st.number_input("Infinium header row", min_value=1, max_value=100,
                            value=detected_inf_header + 1, step=1,
                            key=f"inf_header_{inf_hash[:8]}_{inf_sheet_key}")
        )
        qb_secondary_header_number = (
            int(st.number_input(
                "QuickBooks secondary header row", min_value=1, max_value=100,
                value=int(detected_qb_secondary_header) + 1, step=1,
                key=f"qb_secondary_header_{qb_secondary_hash[:8]}_{qb_secondary_sheet_key}",
            ))
            if detected_qb_secondary_header is not None else None
        )
        inf_secondary_header_number = (
            int(st.number_input(
                "Infinium secondary header row", min_value=1, max_value=100,
                value=int(detected_inf_secondary_header) + 1, step=1,
                key=f"inf_secondary_header_{inf_secondary_hash[:8]}_{inf_secondary_sheet_key}",
            ))
            if detected_inf_secondary_header is not None else None
        )

    try:
        qb_raw = read_source_file(
            qb_bytes, qb_file.name, qb_header_number - 1, qb_sheet_name
        )
        inf_raw = read_source_file(
            inf_bytes, inf_file.name, inf_header_number - 1, inf_sheet_name
        )
        qb_secondary_raw = (
            read_source_file(
                qb_secondary_bytes,
                qb_secondary_file.name,
                int(qb_secondary_header_number) - 1,
                qb_secondary_sheet_name,
            )
            if qb_secondary_bytes and qb_secondary_file
            and qb_secondary_header_number is not None else None
        )
        inf_secondary_raw = (
            read_source_file(
                inf_secondary_bytes,
                inf_secondary_file.name,
                int(inf_secondary_header_number) - 1,
                inf_secondary_sheet_name,
            )
            if inf_secondary_bytes and inf_secondary_file
            and inf_secondary_header_number is not None else None
        )
    except Exception as exc:
        st.error(f"The application could not read the uploaded files with the selected header rows: {exc}")
        return
    if qb_raw.empty or inf_raw.empty:
        st.error("Both source files must contain at least one nonblank data row beneath the selected header row.")
        return
    if qb_secondary_raw is not None and qb_secondary_raw.empty:
        st.error("The QuickBooks secondary upload contains no usable rows beneath its selected header.")
        return
    if inf_secondary_raw is not None and inf_secondary_raw.empty:
        st.error("The Infinium secondary upload contains no usable rows beneath its selected header.")
        return
    qb_import_audit = dict(qb_raw.attrs.get("ingestion_audit", {}))
    inf_import_audit = dict(inf_raw.attrs.get("ingestion_audit", {}))
    qb_secondary_import_audit = (
        dict(qb_secondary_raw.attrs.get("ingestion_audit", {}))
        if qb_secondary_raw is not None else {}
    )
    inf_secondary_import_audit = (
        dict(inf_secondary_raw.attrs.get("ingestion_audit", {}))
        if inf_secondary_raw is not None else {}
    )

    qb_mapping, inf_mapping = mapping_panel(qb_raw, inf_raw, key_suffix)
    qb_secondary_mapping, inf_secondary_mapping = secondary_mapping_panel(
        qb_secondary_raw, inf_secondary_raw, key_suffix
    )
    st.sidebar.markdown("### Reconciliation settings")
    current_central = datetime.now(CENTRAL_TIMEZONE)
    fiscal_year = int(st.sidebar.number_input(
        "Fiscal year", min_value=1900, max_value=2199,
        value=current_central.year, step=1,
        help="Optional input used only for fiscal period identification. This field never restricts matching.",
        key=f"fiscal_year_{key_suffix}",
    ))
    fiscal_period = st.sidebar.selectbox(
        "Current fiscal period", options=[None, *range(1, 14)],
        index=min(current_central.month, 13),
        format_func=lambda value: "All periods" if value is None else f"Period {value:02d}",
        help=(
            "When selected, the Product Aggregate Summary includes only primary QuickBooks "
            "rows whose first-column period equals this value. Transaction matching remains unrestricted."
        ),
        key=f"fiscal_period_{key_suffix}",
    )

    if fiscal_period is not None and len(qb_raw.columns):
        # The primary QuickBooks first column is the authoritative fiscal-period
        # field for aggregate filtering and exception-period reporting.
        qb_mapping["period"] = str(qb_raw.columns[0])

    required_values = [qb_mapping.get(key) for key in ("po", "invoice", "amount")] + [
        inf_mapping.get(key) for key in ("po", "invoice", "amount", "period")
    ]
    if qb_secondary_mapping is not None:
        required_values.extend(qb_secondary_mapping.get(key) for key in ("po", "invoice", "amount"))
    if inf_secondary_mapping is not None:
        required_values.extend(inf_secondary_mapping.get(key) for key in ("po", "invoice", "amount"))
    mappings_ready = not any(value is None for value in required_values)
    signature = (
        qb_hash, inf_hash, qb_secondary_hash, inf_secondary_hash,
        qb_sheet_name, inf_sheet_name, qb_secondary_sheet_name, inf_secondary_sheet_name,
        qb_header_number, inf_header_number, qb_secondary_header_number, inf_secondary_header_number,
        tuple(qb_mapping.items()), tuple(inf_mapping.items()), fiscal_year, fiscal_period,
        tuple((qb_secondary_mapping or {}).items()),
        tuple((inf_secondary_mapping or {}).items()),
        APP_VERSION, MATCHING_RULE_VERSION,
    )
    clear_results_if_signature_changed(signature)
    reconciliation_complete = "reconciliation_result" in st.session_state
    
    if not mappings_ready:
        with progress_container:
            render_ingestion_flow(True, True, False, reconciliation_complete)
            
        st.error(
            "PO, invoice, and amount mappings are required for every uploaded source. "
            "The primary Infinium fiscal-period column is also required."
        )
        return

    qb_detail, qb_subtotal_rows_excluded = filter_qb_subtotal_rows(qb_raw, qb_mapping)
    qb_filter_audit = dict(qb_detail.attrs.get("qb_subtotal_filter_audit", {}))
    qb_secondary_detail: Optional[pd.DataFrame] = None
    qb_secondary_subtotal_rows_excluded = 0
    if qb_secondary_raw is not None and qb_secondary_mapping is not None:
        qb_secondary_detail, qb_secondary_subtotal_rows_excluded = filter_qb_subtotal_rows(
            qb_secondary_raw, qb_secondary_mapping
        )
    if qb_detail.empty:
        with progress_container:
            render_ingestion_flow(
                True, True, True, reconciliation_complete, source_validation_failed=True
            )
            
        st.error(
            "No QuickBooks transaction-detail rows remain. A retained detail row must be populated in every field other than Quantity and Amount."
        )
        return

    first_period_column = str(qb_detail.columns[0])
    period_populated, period_valid, period_strict = fiscal_period_column_profile(
        qb_detail[first_period_column]
    )
    if fiscal_period is None:
        period_validation = pd.DataFrame([{
            "Dataset": "QuickBooks",
            "Check": "First-column aggregate fiscal period",
            "Status": "PASS",
            "Details": "No current fiscal period was selected; Product Aggregate Summary includes all primary QuickBooks periods.",
        }])
    else:
        selected_period_rows = int(
            pd.to_numeric(qb_detail[first_period_column], errors="coerce")
            .eq(int(fiscal_period))
            .sum()
        )
        period_validation = pd.DataFrame([{
            "Dataset": "QuickBooks",
            "Check": "First-column aggregate fiscal period",
            "Status": "FAIL" if not period_strict else ("WARNING" if selected_period_rows == 0 else "PASS"),
            "Details": (
                f"{period_valid:,} of {period_populated:,} populated first-column value(s) are valid periods 1-13; "
                f"{selected_period_rows:,} row(s) belong to selected Period {int(fiscal_period):02d}."
            ),
        }])

    validation_frames = [
        build_source_validation_report(qb_detail, qb_mapping, "QB", qb_import_audit),
        period_validation,
        build_source_validation_report(inf_raw, inf_mapping, "INF", inf_import_audit),
    ]
    if qb_secondary_detail is not None and qb_secondary_mapping is not None:
        validation_frames.append(build_source_validation_report(
            qb_secondary_detail,
            qb_secondary_mapping,
            "QB",
            qb_secondary_import_audit,
            require_period=False,
            dataset_label="QuickBooks Secondary (Historical)",
        ))
    if inf_secondary_raw is not None and inf_secondary_mapping is not None:
        validation_frames.append(build_source_validation_report(
            inf_secondary_raw,
            inf_secondary_mapping,
            "INF",
            inf_secondary_import_audit,
            require_period=False,
            dataset_label="Infinium Secondary (Historical)",
        ))
    validation_report = pd.concat(validation_frames, ignore_index=True)
    validation_failed = validation_report["Status"].eq("FAIL").any()
    validation_warnings = int(validation_report["Status"].eq("WARNING").sum())
    
    with progress_container:
        render_ingestion_flow(
            True, True, True, reconciliation_complete,
            source_validation_failed=validation_failed,
        )
        
    if validation_failed:
        st.error(
            "Source validation failed. Correct the selected header row, required mappings, or malformed source data before matching."
        )
    elif validation_warnings:
        st.info(
            f"Source validation passed with {validation_warnings:,} review warning(s). Harmless blank columns and repeated header rows were ignored."
        )
    else:
        show_toast_once(
            f"source_validation_{key_suffix}_{qb_header_number}_{inf_header_number}",
            "Source validation passed. Both datasets are ready for reconciliation.",
        )
    with st.expander("Source validation details", expanded=validation_failed):
        st.dataframe(validation_report, use_container_width=True, hide_index=True)
    if validation_failed:
        return

    render_pre_execution_controls(qb_detail, inf_raw, qb_mapping, inf_mapping)
    if qb_secondary_detail is not None or inf_secondary_raw is not None:
        st.caption(
            "Secondary historical scope: "
            f"{len(qb_secondary_detail) if qb_secondary_detail is not None else 0:,} QuickBooks row(s) may clear only primary Infinium exceptions; "
            f"{len(inf_secondary_raw) if inf_secondary_raw is not None else 0:,} Infinium row(s) may clear only primary QuickBooks exceptions. "
            "Unused historical rows are ignored."
        )
    if qb_subtotal_rows_excluded:
        st.caption(
            f"Excluded {qb_subtotal_rows_excluded:,} QuickBooks incomplete or subtotal row(s). "
            "Every retained QuickBooks detail row is populated in all fields other than Quantity and Amount."
        )

    _, run_button_col, _ = st.columns([1.25, 1, 1.25])
    with run_button_col:
        run_requested = st.button("Run Reconciliation", type="primary", use_container_width=True)
    if run_requested:
        run_id, run_timestamp = run_id_for_inputs(
            qb_hash, inf_hash, qb_secondary_hash, inf_secondary_hash
        )
        metadata = {
            "run_id": run_id,
            "run_timestamp_dt": run_timestamp,
            "run_timestamp": format_central_timestamp(run_timestamp),
            "qb_filename": qb_file.name,
            "qb_sheet_name": qb_sheet_name,
            "qb_sha256": qb_hash,
            "inf_filename": inf_file.name,
            "inf_sheet_name": inf_sheet_name,
            "inf_sha256": inf_hash,
            "qb_secondary_filename": qb_secondary_file.name if qb_secondary_file else None,
            "qb_secondary_sheet_name": qb_secondary_sheet_name,
            "qb_secondary_sha256": qb_secondary_hash if qb_secondary_file else None,
            "inf_secondary_filename": inf_secondary_file.name if inf_secondary_file else None,
            "inf_secondary_sheet_name": inf_secondary_sheet_name,
            "inf_secondary_sha256": inf_secondary_hash if inf_secondary_file else None,
            "fiscal_year": fiscal_year,
            "fiscal_period": fiscal_period,
            "qb_subtotal_rows_excluded": qb_subtotal_rows_excluded,
            "qb_subtotal_filter_audit": qb_filter_audit,
            "qb_import_audit": qb_import_audit,
            "inf_import_audit": inf_import_audit,
            "qb_secondary_import_audit": qb_secondary_import_audit,
            "inf_secondary_import_audit": inf_secondary_import_audit,
            "qb_secondary_subtotal_rows_excluded": qb_secondary_subtotal_rows_excluded,
            "source_validation_warnings": validation_warnings,
        }
        try:
            with st.spinner("Reconciling source rows and validating accounting controls..."):
                result = build_reconciliation(
                    qb_detail, inf_raw, qb_mapping, inf_mapping, metadata,
                    fiscal_year,
                    qb_secondary_detail,
                    inf_secondary_raw,
                    qb_secondary_mapping,
                    inf_secondary_mapping,
                )
            st.session_state.reconciliation_result = result
            # Workbook bytes are generated only from the Downloads tab. This
            # keeps the reconciliation action focused on matching and controls.
            st.session_state.pop("primary_workbook", None)
            st.session_state.pop("analytics_workbook", None)
            st.rerun()
        except Exception as exc:
            st.error(f"Reconciliation stopped safely: {exc}")
            return

    if "reconciliation_result" in st.session_state:
        render_result(st.session_state.reconciliation_result)


if __name__ == "__main__":
    main()