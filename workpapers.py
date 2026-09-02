"""Excel workpaper construction.

This module is the high-level Excel orchestration layer: it builds the two
downloadable workbooks (the primary accounting workpaper and the optional
analytics evidence package) sheet by sheet. Every function here composes the
styling primitives in ``excel_styles.py`` and pulls its data from a
``ReconciliationResult``. It does not know about Streamlit at all -- the UI
layer (``ui_components.py``) is the only caller that talks to both this
module and the browser.
"""

from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from typing import Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from config import (
    AMBER,
    CENTRAL_TIMEZONE,
    GREEN_LIGHT,
    NAVY,
    NAVY_LIGHT,
    ORANGE,
    RED_LIGHT,
    SLATE,
    SLATE_LIGHT,
    TEAL,
    TEAL_LIGHT,
    TEXT,
    WHITE,
)
from excel_styles import (
    _apply_duplicate_style,
    _apply_number_formats,
    _autofit_workbook_columns,
    _format_body_block,
    _format_header,
    _prepare_sheet,
    _set_widths,
    _thin_border,
    _total_border,
    _write_caption_band,
    _write_title_band,
    _write_total_row,
)
from matching import (
    AMOUNT_CENTS,
    INF_ID,
    QB_ID,
    ReconciliationResult,
    build_fiscal_exception_summary,
    cents_or_zero,
    cents_to_float,
    numeric_quantity_sum,
    numeric_sum,
    valid_cents,
)
from utils import excel_safe, format_central_timestamp


def _write_dataframe_values(ws, frame: pd.DataFrame, start_row: int, start_col: int) -> None:
    for col_offset, header in enumerate(frame.columns):
        ws.cell(start_row, start_col + col_offset, excel_safe(str(header)))
    for row_offset, row in enumerate(frame.values.tolist(), 1):
        for col_offset, value in enumerate(row):
            ws.cell(start_row + row_offset, start_col + col_offset, excel_safe(value))


def _source_totals(frame: pd.DataFrame, mapping: dict[str, Optional[str]]) -> dict[str, float]:
    totals: dict[str, float] = {}
    amount_col = mapping.get("amount")
    quantity_col = mapping.get("quantity")
    if amount_col:
        totals[amount_col] = numeric_sum(frame[amount_col])
    if quantity_col:
        totals[quantity_col] = numeric_quantity_sum(frame[quantity_col])
    return totals


def _duplicate_source_indexes(
    result: ReconciliationResult,
    dataset: str,
) -> set[int]:
    """Return duplicate primary-row indexes, including for pre-2.9 session results."""
    attribute = "duplicate_qb_rows" if dataset == "QuickBooks" else "duplicate_inf_rows"
    stored_indexes = getattr(result, attribute, None)
    if stored_indexes is not None:
        return {int(index) for index in stored_indexes}

    analysis = getattr(result, "duplicate_analysis", pd.DataFrame())
    if analysis.empty or not {"Dataset", "Source Row IDs"}.issubset(analysis.columns):
        return set()
    duplicate_ids: set[str] = set()
    source_rows = analysis.loc[analysis["Dataset"].eq(dataset), "Source Row IDs"]
    for value in source_rows.dropna().astype(str):
        duplicate_ids.update(item.strip() for item in value.split(";") if item.strip())
    frame = result.qb_work if dataset == "QuickBooks" else result.inf_work
    id_column = QB_ID if dataset == "QuickBooks" else INF_ID
    return {
        int(index)
        for index in frame.index
        if str(frame.at[index, id_column]) in duplicate_ids
    }


def _resolve_paired_records_bulk(result: ReconciliationResult) -> list[dict[str, Any]]:
    """Bulk upgrade paired-row records using native dicts to prevent O(N) DataFrame lookups."""
    resolved_list = [dict(record) for record in result.paired_rows]

    # Current matching results already contain explicit source scopes. Return
    # them immediately and avoid rebuilding information that is already known.
    if all(
        "QB Record Scope" in record and "Infinium Record Scope" in record
        for record in resolved_list
    ):
        return resolved_list

    # Legacy results may require their scopes and indexes to be reconstructed
    # from historical-clearance evidence. A grouped clearance intentionally
    # repeats its Clearance ID across multiple sequences, so Clearance ID alone
    # is not unique; the sequence is part of the lookup key.
    clearances = getattr(result, "historical_clearances", pd.DataFrame())
    clearance_map: dict[tuple[Any, Any], dict[str, Any]] = {}
    clearance_fallback: dict[Any, dict[str, Any]] = {}
    if not clearances.empty and "Clearance ID" in clearances.columns:
        for clearance in clearances.to_dict("records"):
            clearance_id = clearance.get("Clearance ID")
            sequence = clearance.get("Group Sequence", 1)
            clearance_map[(clearance_id, sequence)] = clearance
            clearance_fallback.setdefault(clearance_id, clearance)

    for resolved in resolved_list:
        if "QB Record Scope" in resolved and "Infinium Record Scope" in resolved:
            continue

        resolved["QB Record Scope"] = "Primary" if resolved.get("QB Index") is not None else None
        resolved["Infinium Record Scope"] = (
            "Primary" if resolved.get("Infinium Index") is not None else None
        )
        if resolved.get("Section") == "01 Matched - Historical Clearance":
            clearance_id = resolved.get("Match ID")
            sequence = resolved.get("Group Sequence", 1)
            clearance = clearance_map.get(
                (clearance_id, sequence), clearance_fallback.get(clearance_id)
            )
            if clearance:
                primary_is_qb = clearance["Primary Dataset"] == "QuickBooks Primary"
                qb_index = (
                    clearance["Primary Row Index"]
                    if primary_is_qb else clearance["Secondary Row Index"]
                )
                inf_index = (
                    clearance["Secondary Row Index"]
                    if primary_is_qb else clearance["Primary Row Index"]
                )
                resolved["QB Index"] = (
                    None if qb_index is None or pd.isna(qb_index) else int(qb_index)
                )
                resolved["Infinium Index"] = (
                    None if inf_index is None or pd.isna(inf_index) else int(inf_index)
                )
                resolved["QB Record Scope"] = (
                    ("Primary" if primary_is_qb else "Historical")
                    if resolved["QB Index"] is not None else None
                )
                resolved["Infinium Record Scope"] = (
                    ("Historical" if primary_is_qb else "Primary")
                    if resolved["Infinium Index"] is not None else None
                )

    return resolved_list


def build_raw_data_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.active
    ws.title = "Raw Data"
    qb_headers = list(result.qb_raw.columns)
    inf_headers = list(result.inf_raw.columns)
    qb_start = 1
    separator_col = len(qb_headers) + 1
    inf_start = separator_col + 1
    header_row = 3
    data_row = 4
    qb_end = len(qb_headers)
    inf_end = inf_start + len(inf_headers) - 1

    _write_title_band(ws, 1, qb_start, qb_end, "QUICKBOOKS | RAW TRANSACTION DETAIL", NAVY)
    _write_title_band(ws, 1, inf_start, inf_end, "INFINIUM | RAW UPLOAD", TEAL)
    _write_caption_band(
        ws, 2, qb_start, qb_end,
        f"{len(result.qb_raw):,} rows | Source control total: ${result.metrics['QuickBooks Source Total']:,.2f} | "
        f"{result.metrics['QuickBooks Subtotal Rows Excluded']:,} subtotal row(s) excluded before matching | "
        f"Generated {format_central_timestamp(result.run_timestamp)}",
        NAVY,
    )
    _write_caption_band(
        ws, 2, inf_start, inf_end,
        f"{len(result.inf_raw):,} rows | Source control total: ${result.metrics['Infinium Source Total']:,.2f} | "
        f"Values preserved before matching | Generated {format_central_timestamp(result.run_timestamp)}",
        TEAL,
    )
    _write_dataframe_values(ws, result.qb_raw, header_row, qb_start)
    _write_dataframe_values(ws, result.inf_raw, header_row, inf_start)
    _format_header(ws, header_row, qb_start, qb_end, NAVY)
    _format_header(ws, header_row, inf_start, inf_end, TEAL)
    _format_body_block(ws, data_row, data_row + len(result.qb_raw) - 1, qb_start, qb_end, NAVY_LIGHT)
    _format_body_block(ws, data_row, data_row + len(result.inf_raw) - 1, inf_start, inf_end, TEAL_LIGHT)
    for source_index in _duplicate_source_indexes(result, "QuickBooks"):
        if 0 <= source_index < len(result.qb_raw):
            _apply_duplicate_style(ws, data_row + source_index, qb_start, qb_end)
    for source_index in _duplicate_source_indexes(result, "Infinium"):
        if 0 <= source_index < len(result.inf_raw):
            _apply_duplicate_style(ws, data_row + source_index, inf_start, inf_end)
    qb_total_row = data_row + len(result.qb_raw)
    inf_total_row = data_row + len(result.inf_raw)
    _write_total_row(ws, qb_total_row, qb_start, qb_end, _source_totals(result.qb_raw, result.qb_mapping), qb_headers, "SOURCE TOTAL")
    _write_total_row(ws, inf_total_row, inf_start, inf_end, _source_totals(result.inf_raw, result.inf_mapping), inf_headers, "SOURCE TOTAL")
    _apply_number_formats(ws, qb_headers, data_row, qb_total_row, qb_start,
                          {result.qb_mapping["amount"]}, {result.qb_mapping.get("quantity") or ""})
    _apply_number_formats(ws, inf_headers, data_row, inf_total_row, inf_start,
                          {result.inf_mapping["amount"]}, set())
    ws.column_dimensions[get_column_letter(separator_col)].width = 3.5
    ws.column_dimensions[get_column_letter(separator_col)].fill = PatternFill("solid", fgColor=WHITE)
    _set_widths(ws, qb_start, qb_end, header_row, qb_total_row)
    _set_widths(ws, inf_start, inf_end, header_row, inf_total_row)
    ws.freeze_panes = f"{get_column_letter(inf_start)}{data_row}"
    ws.print_title_rows = "1:3"
    _prepare_sheet(ws)


def _paired_display_frames(result: ReconciliationResult) -> tuple[pd.DataFrame, pd.DataFrame, list[str]]:
    def ordered_union(primary_headers: list[str], historical_headers: list[str]) -> list[str]:
        output = list(primary_headers)
        output.extend(header for header in historical_headers if header not in output)
        return output

    def unique_context_header(base: str, headers: list[str]) -> str:
        candidate = base
        suffix = 2
        while candidate in headers:
            candidate = f"{base} {suffix}"
            suffix += 1
        return candidate

    paired_records = _resolve_paired_records_bulk(result)
    qb_historical_used = any(
        record.get("QB Record Scope") == "Historical" for record in paired_records
    )
    inf_historical_used = any(
        record.get("Infinium Record Scope") == "Historical" for record in paired_records
    )
    qb_historical_headers = (
        list(result.qb_secondary_raw.columns)
        if qb_historical_used and result.qb_secondary_raw is not None else []
    )
    inf_historical_headers = (
        list(result.inf_secondary_raw.columns)
        if inf_historical_used and result.inf_secondary_raw is not None else []
    )
    qb_headers = ordered_union(list(result.qb_raw.columns), qb_historical_headers)
    inf_headers = ordered_union(list(result.inf_raw.columns), inf_historical_headers)
    qb_context_header = unique_context_header("QuickBooks Record Context", qb_headers)
    inf_context_header = unique_context_header("Infinium Record Context", inf_headers)

    qb_work_dict = result.qb_work.to_dict("index") if result.qb_work is not None else {}
    inf_work_dict = result.inf_work.to_dict("index") if result.inf_work is not None else {}
    qb_sec_dict = result.qb_secondary_work.to_dict("index") if result.qb_secondary_work is not None else {}
    inf_sec_dict = result.inf_secondary_work.to_dict("index") if result.inf_secondary_work is not None else {}

    def row_values(
        index: Optional[int],
        scope: Optional[str],
        primary_dict: dict,
        historical_dict: dict,
        headers: list[str],
    ) -> list[Any]:
        if index is None:
            return [None] * len(headers)
        source_dict = historical_dict if scope == "Historical" else primary_dict
        row_data = source_dict.get(index, {})
        return [row_data.get(header) for header in headers]

    qb_rows, inf_rows, match_results = [], [], []
    for record in paired_records:
        qidx, iidx = record["QB Index"], record["Infinium Index"]
        qb_scope = record.get("QB Record Scope")
        inf_scope = record.get("Infinium Record Scope")
        qb_values = row_values(
            qidx, qb_scope, qb_work_dict, qb_sec_dict, qb_headers
        )
        inf_values = row_values(
            iidx, inf_scope, inf_work_dict, inf_sec_dict, inf_headers
        )
        qb_values.append(
            "QuickBooks Prior Period Match"
            if qb_scope == "Historical" else "Primary QuickBooks Upload" if qidx is not None else None
        )
        inf_values.append(
            "Infinium Prior Period Match"
            if inf_scope == "Historical" else "Primary Infinium Upload" if iidx is not None else None
        )
        qb_rows.append(qb_values)
        inf_rows.append(inf_values)
        match_results.append(record["Match Result"])
    return (
        pd.DataFrame(qb_rows, columns=qb_headers + [qb_context_header]),
        pd.DataFrame(inf_rows, columns=inf_headers + [inf_context_header]),
        match_results,
    )


def build_reconciled_data_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Reconciled Data")
    qb_display, inf_display, match_results = _paired_display_frames(result)
    qb_headers = list(qb_display.columns)
    inf_headers = list(inf_display.columns)
    qb_start = 1
    match_col = len(qb_headers) + 1
    inf_start = match_col + 1
    qb_end = len(qb_headers)
    inf_end = inf_start + len(inf_headers) - 1
    header_row, data_row = 3, 4
    final_data_row = data_row + len(match_results) - 1

    _write_title_band(ws, 1, qb_start, qb_end, "QUICKBOOKS | RECONCILED", NAVY)
    _write_title_band(ws, 1, match_col, match_col, "MATCH RESULT", SLATE)
    _write_title_band(ws, 1, inf_start, inf_end, "INFINIUM | RECONCILED", TEAL)
    _write_caption_band(
        ws, 2, qb_start, qb_end,
        f"Every primary QuickBooks record appears once. Any accepted QuickBooks prior-period match is displayed on this side and labeled in Record Context. Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    _write_caption_band(ws, 2, match_col, match_col, "Matching Methodology", SLATE)
    _write_caption_band(
        ws, 2, inf_start, inf_end,
        "Every primary Infinium record appears once. Accepted prior-period matches are displayed; unused historical rows are excluded.",
        TEAL,
    )
    _write_dataframe_values(ws, qb_display, header_row, qb_start)
    ws.cell(header_row, match_col, "Match Result")
    for offset, value in enumerate(match_results, 1):
        ws.cell(header_row + offset, match_col, value)
    _write_dataframe_values(ws, inf_display, header_row, inf_start)
    _format_header(ws, header_row, qb_start, qb_end, NAVY)
    _format_header(ws, header_row, match_col, match_col, SLATE)
    _format_header(ws, header_row, inf_start, inf_end, TEAL)
    _format_body_block(ws, data_row, final_data_row, qb_start, qb_end, NAVY_LIGHT)
    _format_body_block(ws, data_row, final_data_row, match_col, match_col, SLATE_LIGHT)
    _format_body_block(ws, data_row, final_data_row, inf_start, inf_end, TEAL_LIGHT)

    duplicate_qb_rows = _duplicate_source_indexes(result, "QuickBooks")
    duplicate_inf_rows = _duplicate_source_indexes(result, "Infinium")
    resolved_records = _resolve_paired_records_bulk(result)
    for offset, record in enumerate(resolved_records):
        row = data_row + offset
        status = str(record["Section"])
        if status == "02 Unmatched QuickBooks":
            for col in range(qb_start, match_col + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=AMBER)
        elif status == "03 Unmatched Infinium":
            for col in range(match_col, inf_end + 1):
                ws.cell(row, col).fill = PatternFill("solid", fgColor=ORANGE)
        qidx = record["QB Index"]
        iidx = record["Infinium Index"]
        if (
            record.get("QB Record Scope") == "Primary"
            and qidx is not None
            and int(qidx) in duplicate_qb_rows
        ):
            _apply_duplicate_style(ws, row, qb_start, qb_end)
        if (
            record.get("Infinium Record Scope") == "Primary"
            and iidx is not None
            and int(iidx) in duplicate_inf_rows
        ):
            _apply_duplicate_style(ws, row, inf_start, inf_end)
        if record.get("QB Record Scope") == "Historical":
            ws.cell(row, qb_end).font = Font(
                name="Segoe UI", size=10, bold=True, color=NAVY
            )
        if record.get("Infinium Record Scope") == "Historical":
            ws.cell(row, inf_end).font = Font(
                name="Segoe UI", size=10, bold=True, color=TEAL
            )
        ws.cell(row, match_col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 30 if "group-level" in str(record["Match Result"]) else 21

    total_row = final_data_row + 1
    _write_total_row(ws, total_row, qb_start, qb_end,
                     _source_totals(result.qb_raw, result.qb_mapping), qb_headers, "RECONCILED TOTAL")
    _write_total_row(ws, total_row, inf_start, inf_end,
                     _source_totals(result.inf_raw, result.inf_mapping), inf_headers, "RECONCILED TOTAL")
    ws.cell(total_row, match_col, f"Control: {result.metrics['Control Status']}")
    ws.cell(total_row, match_col).fill = PatternFill("solid", fgColor=GREEN_LIGHT if result.metrics["Control Status"] == "PASS" else RED_LIGHT)
    ws.cell(total_row, match_col).font = Font(name="Segoe UI", size=10, bold=True, color=TEXT)
    ws.cell(total_row, match_col).border = _total_border()
    ws.cell(total_row, match_col).alignment = Alignment(horizontal="center", vertical="center")
    _apply_number_formats(ws, qb_headers, data_row, total_row, qb_start,
                          {result.qb_mapping["amount"]}, {result.qb_mapping.get("quantity") or ""})
    _apply_number_formats(ws, inf_headers, data_row, total_row, inf_start,
                          {result.inf_mapping["amount"]}, set())
    _set_widths(ws, qb_start, qb_end, header_row, total_row)
    _set_widths(ws, inf_start, inf_end, header_row, total_row)
    ws.column_dimensions[get_column_letter(match_col)].width = 43
    ws.freeze_panes = f"{get_column_letter(inf_start)}{data_row}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(inf_end)}{final_data_row}"
    ws.print_title_rows = "1:3"
    _prepare_sheet(ws)


def build_unresolved_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Unresolved Exceptions")
    source_headers = list(result.qb_raw.columns)
    headers = source_headers + ["Exception Status", "Reference Amount Difference", "Reviewer Note"]
    candidate_map = result.candidates.set_index("QuickBooks Row ID").to_dict("index") if not result.candidates.empty else {}
    
    qb_subset_dict = result.qb_work.loc[result.unmatched_qb].to_dict("index")
    records = []
    for qidx in result.unmatched_qb:
        row_data = qb_subset_dict[qidx]
        candidate = candidate_map.get(row_data[QB_ID], {})
        records.append(
            [row_data.get(col) for col in source_headers]
            + [
                candidate.get("Disposition", "Unmatched QuickBooks"),
                candidate.get("Minimum Amount Difference"),
                "",
            ]
        )
    frame = pd.DataFrame(records, columns=headers)
    end_col = len(headers)
    amount_col_position = source_headers.index(result.qb_mapping["amount"]) + 1
    
    unmatched_qb_amounts = result.qb_work.loc[result.unmatched_qb, AMOUNT_CENTS].tolist()
    amounts = [cents_or_zero(val) for val in unmatched_qb_amounts]
    debits = sum(value for value in amounts if value > 0)
    credits = abs(sum(value for value in amounts if value < 0))
    net = sum(amounts)

    _write_title_band(ws, 1, 1, end_col, "ACCOUNTING EXCEPTIONS | JOURNAL ENTRY SUPPORT", NAVY)
    invalid_unresolved = sum(
        1 for idx in result.unmatched_qb if not valid_cents(result.qb_work.at[idx, AMOUNT_CENTS])
    )
    _write_caption_band(
        ws, 2, 1, end_col,
        f"Net signed support total excludes {invalid_unresolved} row(s) with invalid or missing amounts. "
        f"Review every exception before posting. Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    kpis = [
        ("Unresolved rows", len(result.unmatched_qb), '#,##0'),
        ("Gross debits", cents_to_float(debits), '$#,##0.00;[Red]($#,##0.00);-'),
        ("Credits", cents_to_float(credits), '$#,##0.00;[Red]($#,##0.00);-'),
        ("Proposed JE support total", cents_to_float(net), '$#,##0.00;[Red]($#,##0.00);-'),
    ]
    for idx, (label, value, number_format) in enumerate(kpis):
        start = 1 + idx * 2
        if start > end_col:
            break
        ws.cell(3, start, label)
        ws.cell(4, start, value)
        ws.cell(3, start).font = Font(name="Segoe UI", size=9, bold=True, color=SLATE)
        ws.cell(4, start).font = Font(name="Segoe UI", size=12, bold=True, color=NAVY)
        ws.cell(4, start).number_format = number_format
        for row in (3, 4):
            ws.cell(row, start).fill = PatternFill("solid", fgColor=SLATE_LIGHT)
            ws.cell(row, start).border = _thin_border()

    header_row, data_row = 6, 7
    _write_dataframe_values(ws, frame, header_row, 1)
    _format_header(ws, header_row, 1, end_col, NAVY)
    if len(frame):
        _format_body_block(ws, data_row, data_row + len(frame) - 1, 1, end_col, NAVY_LIGHT)
        duplicate_qb_rows = _duplicate_source_indexes(result, "QuickBooks")
        for offset, qidx in enumerate(result.unmatched_qb):
            row = data_row + offset
            ws.cell(row, source_headers.index(result.qb_mapping["amount"]) + 1).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            ws.cell(row, len(source_headers) + 1).fill = PatternFill("solid", fgColor=AMBER)
            ws.cell(row, len(source_headers) + 1).alignment = Alignment(wrap_text=True, vertical="center")
            ws.cell(row, len(source_headers) + 2).number_format = '$#,##0.00;[Red]($#,##0.00);-'
            if int(qidx) in duplicate_qb_rows:
                _apply_duplicate_style(ws, row, 1, end_col)
    total_row = data_row + len(frame)
    _write_total_row(
        ws, total_row, 1, end_col,
        {result.qb_mapping["amount"]: cents_to_float(net)}, headers,
        "PROPOSED JE SUPPORT TOTAL",
    )
    ws.cell(total_row, amount_col_position).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    ws.column_dimensions[get_column_letter(len(source_headers) + 1)].width = 48
    ws.column_dimensions[get_column_letter(len(source_headers) + 2)].width = 24
    ws.column_dimensions[get_column_letter(len(source_headers) + 3)].width = 36
    _set_widths(ws, 1, len(source_headers), header_row, total_row)
    ws.freeze_panes = f"A{data_row}"
    if len(frame):
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(end_col)}{header_row + len(frame)}"
        note_col = get_column_letter(len(source_headers) + 3)
        validation = DataValidation(
            type="textLength", operator="lessThanOrEqual", formula1="1000", allow_blank=True
        )
        validation.error = "Reviewer notes are limited to 1,000 characters."
        validation.errorTitle = "Note too long"
        ws.add_data_validation(validation)
        validation.add(f"{note_col}{data_row}:{note_col}{header_row + len(frame)}")

    journal_amount = abs(cents_to_float(net))
    je_title_row = total_row + 3
    je_caption_row = je_title_row + 1
    je_header_row = je_title_row + 2
    je_data_row = je_header_row + 1
    je_headers = [
        "Entry Name", "GL Account", "Account Name", "Debit", "Credit", "Entry Basis",
    ]
    je_frame = pd.DataFrame(
        [
            [
                "AC001 Sales Accrual",
                "017-00000-110160.0",
                "Accrued Income",
                journal_amount,
                0.0,
                "Unresolved QuickBooks net exception support",
            ],
            [
                "AC001 Sales Accrual",
                "017-91000-400000-0",
                "Income-Manufacturing",
                0.0,
                journal_amount,
                "Balanced offset",
            ],
        ],
        columns=je_headers,
    )
    _write_title_band(
        ws, je_title_row, 1, end_col,
        "PROPOSED JOURNAL ENTRY | AC001 SALES ACCRUAL",
        SLATE,
    )
    _write_caption_band(
        ws, je_caption_row, 1, end_col,
        "Post only after review and approval. Debit 017-00000-110160.0 Accrued Income and credit "
        "017-91000-400000-0 Income-Manufacturing for the absolute unresolved net amount; evaluate "
        "reversals and negative source values before posting.",
        SLATE,
    )
    _write_dataframe_values(ws, je_frame, je_header_row, 1)
    _format_header(ws, je_header_row, 1, len(je_headers), SLATE)
    _format_body_block(ws, je_data_row, je_data_row + len(je_frame) - 1, 1, len(je_headers), SLATE_LIGHT)
    for row in range(je_data_row, je_data_row + len(je_frame)):
        ws.cell(row, 4).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        ws.cell(row, 5).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    je_total_row = je_data_row + len(je_frame)
    _write_total_row(
        ws, je_total_row, 1, len(je_headers),
        {"Debit": journal_amount, "Credit": journal_amount}, je_headers, "BALANCED TOTAL",
    )
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 0, 24)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 23)
    ws.column_dimensions["C"].width = max(ws.column_dimensions["C"].width or 0, 28)
    ws.column_dimensions["F"].width = max(ws.column_dimensions["F"].width or 0, 52)

    duplicate_frame = result.duplicate_analysis
    duplicate_headers = list(duplicate_frame.columns)
    duplicate_title_row = je_total_row + 3
    duplicate_header_row = duplicate_title_row + 2
    duplicate_data_row = duplicate_header_row + 1
    duplicate_end_col = len(duplicate_headers)
    section_end_col = max(end_col, duplicate_end_col)
    _write_title_band(
        ws, duplicate_title_row, 1, section_end_col,
        "DUPLICATE MATCHING-KEY REVIEW | QUICKBOOKS AND INFINIUM", TEAL,
    )
    duplicate_caption = (
        f"{len(duplicate_frame):,} repeated matching-key group(s) identified. Review these source values for "
        "duplicate entries or legitimate repeated transactions."
        if len(duplicate_frame)
        else "No repeated PO/invoice/amount matching keys were identified in either dataset."
    )
    _write_caption_band(ws, duplicate_title_row + 1, 1, section_end_col, duplicate_caption, TEAL)
    _write_dataframe_values(ws, duplicate_frame, duplicate_header_row, 1)
    _format_header(ws, duplicate_header_row, 1, duplicate_end_col, TEAL)
    if len(duplicate_frame):
        duplicate_last_row = duplicate_data_row + len(duplicate_frame) - 1
        _format_body_block(ws, duplicate_data_row, duplicate_last_row, 1, duplicate_end_col, TEAL_LIGHT)
        _apply_number_formats(
            ws, duplicate_headers, duplicate_data_row, duplicate_last_row, 1,
            {"Amount"}, set(),
        )
        for row in range(duplicate_data_row, duplicate_last_row + 1):
            _apply_duplicate_style(ws, row, 1, duplicate_end_col)
        ws.column_dimensions[get_column_letter(duplicate_headers.index("Source Row IDs") + 1)].width = 44
        ws.column_dimensions[get_column_letter(duplicate_headers.index("Reconciliation Status") + 1)].width = 42
    else:
        duplicate_last_row = duplicate_header_row

    fiscal_summary = build_fiscal_exception_summary(result)
    fiscal_headers = list(fiscal_summary.columns)
    fiscal_title_row = duplicate_last_row + 3
    fiscal_caption_row = fiscal_title_row + 1
    fiscal_header_row = fiscal_title_row + 2
    fiscal_data_row = fiscal_header_row + 1
    fiscal_end_col = len(fiscal_headers)
    fiscal_section_end_col = max(end_col, fiscal_end_col)
    selected_period = result.metadata.get("fiscal_period")
    has_fiscal_period = bool(result.qb_mapping.get("period"))
    _write_title_band(
        ws, fiscal_title_row, 1, fiscal_section_end_col,
        (
            "FISCAL-PERIOD EXCEPTION SUMMARY | CURRENT VS PRIOR PERIODS"
            if has_fiscal_period
            else "EXCEPTION SUMMARY | FISCAL PERIOD NOT AVAILABLE"
        ),
        NAVY,
    )
    quantity_note = (
        "Exception quantity is sourced from the mapped QuickBooks quantity column."
        if result.qb_mapping.get("quantity")
        else "No QuickBooks quantity column was mapped; exception quantities are shown as zero."
    )
    _write_caption_band(
        ws, fiscal_caption_row, 1, fiscal_section_end_col,
        (
            f"Selected current reporting period: PD-{int(selected_period):02d}. Every other valid QuickBooks fiscal "
            f"period is classified as a prior-period urgent exception. {quantity_note}"
            if has_fiscal_period and selected_period is not None
            else f"No current reporting period was selected. Exceptions are summarized by source period without current/prior classification. {quantity_note}"
            if has_fiscal_period
            else "No credible QuickBooks fiscal-period identifier was found or mapped. Period-based "
            f"classification is disabled and all exceptions are summarized together. {quantity_note}"
        ),
        NAVY,
    )
    _write_dataframe_values(ws, fiscal_summary, fiscal_header_row, 1)
    _format_header(ws, fiscal_header_row, 1, fiscal_end_col, NAVY)
    if len(fiscal_summary):
        fiscal_last_row = fiscal_data_row + len(fiscal_summary) - 1
        _format_body_block(ws, fiscal_data_row, fiscal_last_row, 1, fiscal_end_col, NAVY_LIGHT)
        for offset, classification in enumerate(fiscal_summary["Period Classification"], start=fiscal_data_row):
            fill = (
                RED_LIGHT
                if classification == "Prior-Period Urgent Exception"
                else GREEN_LIGHT
                if classification == "Current Reporting Period"
                else AMBER
            )
            for col in range(1, fiscal_end_col + 1):
                ws.cell(offset, col).fill = PatternFill("solid", fgColor=fill)
        _apply_number_formats(
            ws, fiscal_headers, fiscal_data_row, fiscal_last_row, 1,
            {"Net Exception Amount"}, {"Exception Count", "Exception Quantity"},
        )
    else:
        fiscal_last_row = fiscal_header_row
    fiscal_total_row = fiscal_last_row + 1
    _write_total_row(
        ws, fiscal_total_row, 1, fiscal_end_col,
        {
            "Exception Count": float(fiscal_summary["Exception Count"].sum()) if len(fiscal_summary) else 0,
            "Exception Quantity": float(fiscal_summary["Exception Quantity"].sum()) if len(fiscal_summary) else 0,
            "Net Exception Amount": float(fiscal_summary["Net Exception Amount"].sum()) if len(fiscal_summary) else 0,
        },
        fiscal_headers,
        "TOTAL EXCEPTIONS",
    )
    _set_widths(ws, 1, fiscal_end_col, fiscal_header_row, fiscal_total_row)
    ws.column_dimensions["B"].width = max(ws.column_dimensions["B"].width or 0, 34)
    ws.print_title_rows = "1:6"
    _prepare_sheet(ws)


def build_product_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.create_sheet("Product Aggregate Summary")
    frame = result.product_summary
    headers = ["Product Name", "Product Quantity", "Product Value"]
    _write_title_band(ws, 1, 1, 3, "PRODUCT AGGREGATE SUMMARY", NAVY)
    selected_period = result.metadata.get("fiscal_period")
    period_scope = (
        f"Only primary QuickBooks rows from Period {int(selected_period):02d} are included."
        if selected_period is not None
        else "All primary QuickBooks fiscal periods are included."
    )
    _write_caption_band(
        ws, 2, 1, 3,
        "Products are classified from the QuickBooks description column. "
        f"{period_scope} This classification does not influence data matching. "
        f"Generated {format_central_timestamp(result.run_timestamp)}.",
        NAVY,
    )
    _write_dataframe_values(ws, frame.reindex(columns=headers), 3, 1)
    _format_header(ws, 3, 1, 3, NAVY)
    if len(frame):
        _format_body_block(ws, 4, 3 + len(frame), 1, 3, NAVY_LIGHT)
    total_row = 4 + len(frame)
    totals = {
        "Product Quantity": float(frame["Product Quantity"].sum()) if not frame.empty else 0,
        "Product Value": float(frame["Product Value"].sum()) if not frame.empty else 0,
    }
    _write_total_row(ws, total_row, 1, 3, totals, headers)
    _apply_number_formats(ws, headers, 4, total_row, 1, {"Product Value"}, {"Product Quantity"})
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.freeze_panes = "A4"
    if len(frame):
        ws.auto_filter.ref = f"A3:C{3 + len(frame)}"
    _prepare_sheet(ws, landscape=False)


def _save_workbook_bytes(wb: Workbook) -> bytes:
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    _autofit_workbook_columns(wb)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _apply_workbook_run_metadata(wb: Workbook, result: ReconciliationResult) -> None:
    central_timestamp = result.run_timestamp.astimezone(CENTRAL_TIMEZONE)
    excel_timestamp = central_timestamp.replace(tzinfo=None)
    display_timestamp = format_central_timestamp(central_timestamp)
    # Visible report captions carry the controlled Central timestamp. Core
    # properties use the same run time, while page headers/footers are blank so
    # Excel cannot surface a stale or locale-generated tag on printed sheets.
    wb.properties.created = excel_timestamp
    wb.properties.modified = excel_timestamp
    existing_description = wb.properties.description or ""
    wb.properties.description = (
        f"{existing_description} Generated {display_timestamp}. Run ID: {result.run_id}."
    ).strip()
    for ws in wb.worksheets:
        for section in (
            ws.oddHeader, ws.evenHeader, ws.firstHeader,
            ws.oddFooter, ws.evenFooter, ws.firstFooter,
        ):
            section.left.text = None
            section.center.text = None
            section.right.text = None


def build_primary_workbook(result: ReconciliationResult) -> bytes:
    wb = Workbook()
    wb.properties.creator = "Sales Reconciliation Application"
    wb.properties.title = f"Sales Reconciliation {result.run_id}"
    wb.properties.subject = "QuickBooks to Infinium reconciliation and journal-entry support"
    wb.properties.description = "Four-sheet accounting workpaper generated from one controlled reconciliation run."
    build_raw_data_sheet(wb, result)
    build_reconciled_data_sheet(wb, result)
    build_unresolved_sheet(wb, result)
    build_product_sheet(wb, result)
    _apply_workbook_run_metadata(wb, result)
    return _save_workbook_bytes(wb)


def detailed_ledger_dataframe(result: ReconciliationResult) -> pd.DataFrame:
    records: list[dict[str, Any]] = []
    qb_display, inf_display, _ = _paired_display_frames(result)
    resolved_records = _resolve_paired_records_bulk(result)
    
    qb_disp_dicts = qb_display.to_dict("records")
    inf_disp_dicts = inf_display.to_dict("records")
    
    qb_id_map = result.qb_work[QB_ID].to_dict() if result.qb_work is not None else {}
    inf_id_map = result.inf_work[INF_ID].to_dict() if result.inf_work is not None else {}
    qb_sec_id_map = result.qb_secondary_work[QB_ID].to_dict() if result.qb_secondary_work is not None else {}
    inf_sec_id_map = result.inf_secondary_work[INF_ID].to_dict() if result.inf_secondary_work is not None else {}

    for position, record in enumerate(resolved_records):
        output = {
            "Run ID": result.run_id,
            "Section": record["Section"],
            "Match ID": record["Match ID"],
            "Match Result": record["Match Result"],
            "Confidence": record["Confidence"],
            "Group Sequence": record["Group Sequence"],
            "Assessment Explanation": record["Explanation"],
        }
        qidx, iidx = record["QB Index"], record["Infinium Index"]
        qb_scope = record.get("QB Record Scope")
        inf_scope = record.get("Infinium Record Scope")
        
        active_q_map = qb_sec_id_map if qb_scope == "Historical" else qb_id_map
        output["QB | Source Row ID"] = active_q_map.get(qidx) if qidx is not None else None
        
        for header, value in qb_disp_dicts[position].items():
            output[f"QB | {header}"] = value
            
        active_i_map = inf_sec_id_map if inf_scope == "Historical" else inf_id_map
        output["INF | Source Row ID"] = active_i_map.get(iidx) if iidx is not None else None
        
        for header, value in inf_disp_dicts[position].items():
            output[f"INF | {header}"] = value
            
        records.append(output)
    return pd.DataFrame(records)


def _add_standard_data_sheet(
    wb: Workbook,
    name: str,
    title: str,
    caption: str,
    frame: pd.DataFrame,
    header_color: str = NAVY,
    chart_column: Optional[str] = None,
) -> Any:
    ws = wb.create_sheet(name)
    end_col = max(len(frame.columns), 1)
    _write_title_band(ws, 1, 1, end_col, title, header_color)
    _write_caption_band(ws, 2, 1, end_col, caption, header_color)
    if frame.empty and len(frame.columns) == 0:
        frame = pd.DataFrame({"Result": ["No records"]})
        end_col = 1
    _write_dataframe_values(ws, frame, 3, 1)
    _format_header(ws, 3, 1, len(frame.columns), header_color)
    if len(frame):
        _format_body_block(ws, 4, 3 + len(frame), 1, len(frame.columns), NAVY_LIGHT)
        _apply_number_formats(ws, list(frame.columns), 4, 3 + len(frame), 1, set(), set())
    _set_widths(ws, 1, len(frame.columns), 3, max(3 + len(frame), 3), maximum=48)
    wrap_terms = ("EXPLANATION", "REQUIREMENT", "DISPOSITION", "MATCH RESULT", "CANDIDATE IDS")
    for col, header in enumerate(frame.columns, 1):
        if any(term in str(header).upper() for term in wrap_terms):
            ws.column_dimensions[get_column_letter(col)].width = 42
            for row in range(4, 4 + len(frame)):
                ws.cell(row, col).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                ws.row_dimensions[row].height = max(ws.row_dimensions[row].height or 20, 32)
    ws.freeze_panes = "A4"
    if len(frame):
        ws.auto_filter.ref = f"A3:{get_column_letter(len(frame.columns))}{3 + len(frame)}"
    if chart_column and chart_column in frame.columns and len(frame):
        data_col = list(frame.columns).index(chart_column) + 1
        category_col = 1
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = f"{chart_column} by Match Method"
        chart.y_axis.title = "Match Method"
        chart.x_axis.title = chart_column
        chart.height = 7.5
        chart.width = 15
        data = Reference(ws, min_col=data_col, min_row=3, max_row=3 + len(frame))
        categories = Reference(ws, min_col=category_col, min_row=4, max_row=3 + len(frame))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.legend = None
        ws.add_chart(chart, f"{get_column_letter(len(frame.columns) + 2)}3")
    _prepare_sheet(ws)
    return ws


def build_analytics_summary_sheet(wb: Workbook, result: ReconciliationResult) -> None:
    ws = wb.active
    ws.title = "Executive Summary"
    _write_title_band(ws, 1, 1, 8, "RECONCILIATION ANALYTICS | EXECUTIVE SUMMARY", NAVY)
    _write_caption_band(
        ws, 2, 1, 8,
        f"Run ID: {result.run_id} | Generated {format_central_timestamp(result.run_timestamp)} | Control status: {result.metrics['Control Status']}",
        NAVY,
    )
    cards = [
        ("QuickBooks rows", result.metrics["QuickBooks Rows"], '#,##0'),
        ("Infinium rows", result.metrics["Infinium Rows"], '#,##0'),
        ("QB match rate", result.metrics["QuickBooks Match Rate by Row"], '0.0%'),
        ("Unresolved QB amount", result.metrics["Unresolved QuickBooks Amount"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("QB source total", result.metrics["QuickBooks Source Total"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Infinium source total", result.metrics["Infinium Source Total"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Matched amount difference", result.metrics["Matched Amount Difference"], '$#,##0.00;[Red]($#,##0.00);-'),
        ("Control status", result.metrics["Control Status"], '@'),
    ]
    for idx, (label, value, number_format) in enumerate(cards):
        row = 4 if idx < 4 else 7
        col = 1 + (idx % 4) * 2
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
        ws.cell(row, col, label)
        ws.cell(row + 1, col, excel_safe(value))
        for r in (row, row + 1):
            for c in (col, col + 1):
                ws.cell(r, c).fill = PatternFill("solid", fgColor=SLATE_LIGHT)
                ws.cell(r, c).border = _thin_border()
        ws.cell(row, col).font = Font(name="Segoe UI", size=9, bold=True, color=SLATE)
        ws.cell(row + 1, col).font = Font(
            name="Segoe UI", size=13, bold=True,
            color=NAVY if value != "FAIL" else "9C0006",
        )
        ws.cell(row + 1, col).number_format = number_format

    start = 11
    ws.cell(start, 1, "MODEL STATUS")
    ws.cell(start, 2, result.metrics["Control Status"])
    ws.cell(start, 1).font = Font(name="Segoe UI", size=11, bold=True, color=WHITE)
    ws.cell(start, 1).fill = PatternFill("solid", fgColor=SLATE)
    ws.cell(start, 2).font = Font(name="Segoe UI", size=11, bold=True, color=TEXT)
    ws.cell(start, 2).fill = PatternFill("solid", fgColor=GREEN_LIGHT if result.metrics["Control Status"] == "PASS" else RED_LIGHT)
    _write_dataframe_values(ws, result.controls, start + 2, 1)
    _format_header(ws, start + 2, 1, len(result.controls.columns), SLATE)
    _format_body_block(ws, start + 3, start + 2 + len(result.controls), 1, len(result.controls.columns), SLATE_LIGHT)
    _apply_number_formats(ws, list(result.controls.columns), start + 3, start + 2 + len(result.controls), 1, set(), set())
    status_col = list(result.controls.columns).index("Status") + 1
    if len(result.controls):
        green = PatternFill("solid", fgColor=GREEN_LIGHT)
        red = PatternFill("solid", fgColor=RED_LIGHT)
        ws.conditional_formatting.add(
            f"{get_column_letter(status_col)}{start + 3}:{get_column_letter(status_col)}{start + 2 + len(result.controls)}",
            FormulaRule(formula=[f'{get_column_letter(status_col)}{start + 3}="PASS"'], fill=green),
        )
        ws.conditional_formatting.add(
            f"{get_column_letter(status_col)}{start + 3}:{get_column_letter(status_col)}{start + 2 + len(result.controls)}",
            FormulaRule(formula=[f'{get_column_letter(status_col)}{start + 3}="FAIL"'], fill=red),
        )
    for col in range(1, 9):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.freeze_panes = "A3"
    _prepare_sheet(ws)


def build_analytics_workbook(result: ReconciliationResult) -> bytes:
    wb = Workbook()
    wb.properties.creator = "Sales Reconciliation Application"
    wb.properties.title = f"Sales Reconciliation Analytics {result.run_id}"
    wb.properties.subject = "Detailed matching evidence and reconciliation controls"
    build_analytics_summary_sheet(wb, result)
    _add_standard_data_sheet(
        wb, "Match Method Summary", "MATCH METHOD SUMMARY",
        "Distribution of matched and unresolved rows. Percentages use QuickBooks row count as the denominator.",
        result.method_summary, NAVY, chart_column="QuickBooks Rows",
    )
    _add_standard_data_sheet(
        wb, "Detailed Match Ledger", "DETAILED MATCH Ledger",
        "One record per displayed reconciliation line. Historical clearances show the specific accepted prior-period row on the opposing side; unused historical rows are excluded.",
        detailed_ledger_dataframe(result), SLATE,
    )
    _add_standard_data_sheet(
        wb, "Normalization Detail", "NORMALIZATION DETAIL",
        "Original source values and the exact normalized references and signed amounts considered by the matching engine.",
        result.normalization, TEAL,
    )
    _add_standard_data_sheet(
        wb, "Match Assessment", "MATCH ASSESSMENT",
        "One row per accepted match group or unresolved QuickBooks decision, including criteria and evidence.",
        result.assessments, NAVY,
    )
    _add_standard_data_sheet(
        wb, "Historical Clearances", "SECONDARY HISTORICAL CLEARANCES",
        "Accepted strict matches between an unresolved primary row and the opposing historical source. Unused secondary rows are intentionally excluded.",
        result.historical_clearances, TEAL,
    )
    _add_standard_data_sheet(
        wb, "Exception Analysis", "EXCEPTION ANALYSIS",
        "Unresolved population summarized by source period, reason, and source ledger. Period is reporting metadata only.",
        result.exception_analysis, NAVY,
    )
    _add_standard_data_sheet(
        wb, "Rules and Run Config", "RULES AND RUN CONFIGURATION",
        "Exact rules, file fingerprints, mappings, versions, and parameters used for this run.",
        pd.concat(
            [
                result.config,
                pd.DataFrame([{"Setting": "", "Value": ""}, {"Setting": "MATCHING RULES", "Value": ""}]),
                result.rules.rename(columns={
                    "Priority": "Setting",
                    "Rule": "Value",
                    "Automatic": "Automatic",
                    "Requirement": "Requirement",
                }),
            ],
            ignore_index=True,
        ),
        SLATE,
    )
    _apply_workbook_run_metadata(wb, result)
    return _save_workbook_bytes(wb)


# Public alias: ui_components.py needs the paired display frames to build the
# in-app "Reconciled View" preview without duplicating this logic.
paired_display_frames = _paired_display_frames
