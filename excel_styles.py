"""OpenPyXL formatting engine for the reconciliation workbooks.

Every visual primitive used by the sheet builders in ``workpapers.py`` lives
here: thin/total borders, header and body block styling, duplicate-row
highlighting, number formats, column-width sizing (including the bounded
workbook-wide autofit pass), title/caption bands, and total rows. Keeping
these out of workpapers.py keeps that module focused on "what goes on each
sheet" rather than "how a cell block is painted."
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import (
    BORDER,
    DUPLICATE_RED_FILL,
    DUPLICATE_RED_TEXT,
    SLATE,
    SLATE_LIGHT,
    TEXT,
    TOTAL_FILL,
    WHITE,
)

# ---------------------------------------------------------------------------
# Global Style Singletons (Prevents massive instantiation overhead in loops)
# ---------------------------------------------------------------------------

ALIGN_LEFT_CENTER = Alignment(horizontal="left", vertical="center")
ALIGN_RIGHT_CENTER = Alignment(horizontal="right", vertical="center")
ALIGN_CENTER_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_WRAP_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

FONT_BODY = Font(name="Segoe UI", size=10, color=TEXT)
FONT_DUPLICATE = Font(name="Segoe UI", size=10, bold=True, color=DUPLICATE_RED_TEXT)
FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color=TEXT)
FONT_HEADER = Font(name="Segoe UI", size=10, bold=True, color=WHITE)
FONT_TITLE = Font(name="Segoe UI", size=12, bold=True, color=WHITE)

FILL_DUPLICATE = PatternFill("solid", fgColor=DUPLICATE_RED_FILL)
FILL_TOTAL = PatternFill("solid", fgColor=TOTAL_FILL)
FILL_NONE = PatternFill(fill_type=None)
FILL_CAPTION_BAND = PatternFill("solid", fgColor=SLATE_LIGHT)

_SIDE_THIN_BORDER = Side(style="thin", color=BORDER)
BORDER_THIN = Border(
    left=_SIDE_THIN_BORDER, right=_SIDE_THIN_BORDER, 
    top=_SIDE_THIN_BORDER, bottom=_SIDE_THIN_BORDER
)

BORDER_TOTAL = Border(
    top=Side(style="thin", color=SLATE),
    bottom=Side(style="double", color=SLATE),
)

# ---------------------------------------------------------------------------
# Formatting Engine
# ---------------------------------------------------------------------------

def _thin_border() -> Border:
    return BORDER_THIN


def _total_border() -> Border:
    return BORDER_TOTAL


def _format_header(ws, row: int, start_col: int, end_col: int, color: str) -> None:
    # Instantiate custom fills and borders once per block, not once per cell
    header_fill = PatternFill("solid", fgColor=color)
    header_border = Border(bottom=Side(style="medium", color=color))
    
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = header_fill
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_WRAP_LEFT
        cell.border = header_border
    ws.row_dimensions[row].height = 34


def _format_body_block(
    ws,
    start_row: int,
    end_row: int,
    start_col: int,
    end_col: int,
    light_fill: str,
) -> None:
    fill_even = PatternFill("solid", fgColor=light_fill)
    
    for row in range(start_row, end_row + 1):
        row_fill = fill_even if row % 2 == 0 else FILL_NONE
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row, col)
            cell.font = FONT_BODY
            cell.border = BORDER_THIN
            cell.fill = row_fill
            cell.alignment = ALIGN_LEFT_CENTER
        ws.row_dimensions[row].height = 20


def _apply_duplicate_style(ws, row: int, start_col: int, end_col: int) -> None:
    """Apply Excel's traditional red bad-value style to a duplicate source row."""
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = FILL_DUPLICATE
        cell.font = FONT_DUPLICATE


def _apply_number_formats(
    ws,
    headers: list[str],
    start_row: int,
    end_row: int,
    start_col: int,
    amount_columns: set[str],
    quantity_columns: set[str],
) -> None:
    for offset, header in enumerate(headers):
        col = start_col + offset
        header_upper = str(header).upper()
        
        # Determine the target alignment and format string for the entire column first
        target_align = None
        target_format = None
        
        if any(term in header_upper for term in ("VALID", "GROUP-LEVEL", "AUTOMATIC")):
            target_align = ALIGN_CENTER_CENTER
        elif any(term in header_upper for term in ("PERCENT", "RATE", "SHARE")):
            target_format = "0.0%"
            target_align = ALIGN_RIGHT_CENTER
        elif any(term in header_upper for term in ("COUNT", "ROWS", "CANDIDATE COUNT")):
            target_format = '#,##0;[Red](#,##0);-'
            target_align = ALIGN_RIGHT_CENTER
        elif header in amount_columns or any(term in header_upper for term in ("AMOUNT", "VALUE", "VARIANCE", "DIFFERENCE", "BALANCE", "EXPOSURE", "TOLERANCE")):
            target_format = '$#,##0.00;[Red]($#,##0.00);-'
            target_align = ALIGN_RIGHT_CENTER
        elif header in quantity_columns or any(term in header_upper for term in ("QUANTITY", "QTY", "COUNT", "ROWS")):
            target_format = '#,##0.00;[Red](#,##0.00);-'
            target_align = ALIGN_RIGHT_CENTER
        elif "DATE" in header_upper or "TIMESTAMP" in header_upper:
            target_format = "yyyy-mm-dd"
            target_align = ALIGN_CENTER_CENTER
            
        # Apply only if a rule matched
        if target_align or target_format:
            for row in range(start_row, end_row + 1):
                cell = ws.cell(row, col)
                if target_align:
                    cell.alignment = target_align
                if target_format:
                    cell.number_format = target_format


def _set_widths(
    ws,
    start_col: int,
    end_col: int,
    start_row: int,
    end_row: int,
    minimum: float = 11,
    maximum: float = 34,
) -> None:
    for col in range(start_col, end_col + 1):
        lengths = [len(str(ws.cell(row, col).value or "")) for row in range(start_row, end_row + 1)]
        ws.column_dimensions[get_column_letter(col)].width = min(max(max(lengths, default=0) + 2, minimum), maximum)


def _autofit_workbook_columns(
    wb: Workbook,
    minimum: float = 10,
    maximum: float = 52,
    sample_rows: int = 750,
) -> None:
    """Auto-size columns from a bounded sample instead of rescanning every cell."""
    for ws in wb.worksheets:
        merged_coordinates: set[str] = set()
        for merged_range in ws.merged_cells.ranges:
            for row in ws.iter_rows(
                min_row=merged_range.min_row,
                max_row=merged_range.max_row,
                min_col=merged_range.min_col,
                max_col=merged_range.max_col,
            ):
                merged_coordinates.update(cell.coordinate for cell in row)

        sampled_row_numbers = list(range(1, min(ws.max_row, sample_rows) + 1))
        if ws.max_row > sample_rows:
            sampled_row_numbers.append(ws.max_row)

        for column_index in range(1, ws.max_column + 1):
            letter = get_column_letter(column_index)
            maximum_length = 0
            has_unmerged_value = False
            for row_index in sampled_row_numbers:
                cell = ws.cell(row_index, column_index)
                if cell.coordinate in merged_coordinates or cell.value is None:
                    continue
                has_unmerged_value = True
                value = cell.value
                if isinstance(value, datetime):
                    display = value.strftime("%Y-%m-%d %I:%M:%S %p")
                elif isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
                    if "$" in str(cell.number_format):
                        display = f"${float(value):,.2f}"
                    elif any(token in str(cell.number_format) for token in ("0.00", "#,##0")):
                        display = f"{float(value):,.2f}"
                    else:
                        display = str(value)
                else:
                    display = str(value)
                maximum_length = max(
                    maximum_length,
                    max((len(line) for line in display.splitlines()), default=0),
                )

            if has_unmerged_value:
                ws.column_dimensions[letter].width = min(max(maximum_length + 2, minimum), maximum)
            else:
                existing = ws.column_dimensions[letter].width or minimum
                ws.column_dimensions[letter].width = min(max(existing, 3.5), maximum)


def _write_title_band(ws, row: int, start_col: int, end_col: int, title: str, color: str) -> None:
    if start_col < end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    
    title_fill = PatternFill("solid", fgColor=color)
    cell = ws.cell(row, start_col, title)
    cell.fill = title_fill
    cell.font = FONT_TITLE
    cell.alignment = ALIGN_LEFT_CENTER
    
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row, col).fill = title_fill
    ws.row_dimensions[row].height = 27


def _write_caption_band(ws, row: int, start_col: int, end_col: int, caption: str, color: str) -> None:
    if start_col < end_col:
        ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
        
    caption_font = Font(name="Segoe UI", size=9, italic=True, color=color)
    cell = ws.cell(row, start_col, caption)
    cell.fill = FILL_CAPTION_BAND
    cell.font = caption_font
    cell.alignment = ALIGN_WRAP_LEFT
    
    for col in range(start_col + 1, end_col + 1):
        ws.cell(row, col).fill = FILL_CAPTION_BAND
    ws.row_dimensions[row].height = 30


def _write_total_row(
    ws,
    row: int,
    start_col: int,
    end_col: int,
    totals: dict[str, float],
    headers: list[str],
    label: str = "TOTAL",
) -> None:
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = FILL_TOTAL
        cell.font = FONT_TOTAL
        cell.border = BORDER_TOTAL
        
    ws.cell(row, start_col, label)
    for offset, header in enumerate(headers):
        if header in totals:
            cell = ws.cell(row, start_col + offset, totals[header])
            cell.number_format = '$#,##0.00;[Red]($#,##0.00);-' if "amount" in header.lower() or "value" in header.lower() else '#,##0.00;[Red](#,##0.00);-'
            cell.alignment = ALIGN_RIGHT_CENTER
    ws.row_dimensions[row].height = 23


def _prepare_sheet(ws, landscape: bool = True) -> None:
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3 if landscape else ws.PAPERSIZE_LETTER
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.autoPageBreaks = False