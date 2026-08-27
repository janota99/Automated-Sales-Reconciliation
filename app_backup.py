import io
import time
from datetime import datetime

import pandas as pd
import streamlit as st

# Openpyxl Core Styling & Calculation Engines
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# MOCK ENGINES (Replace these with your actual local imports if available)
try:
    from engine.matching import execute_sales_reconciliation, get_fuzzy_lexicon_match
except ImportError:
    def execute_sales_reconciliation(df_qb, df_inf, qb_cols, inf_cols):
        df_qb = df_qb.copy()
        df_qb["IS_MATCHED"] = False
        df_qb["MATCH_METHOD"] = "Unmatched / No Match"
        return df_qb, df_inf

    def get_fuzzy_lexicon_match(memo):
        return "Standard Product Line" if pd.notna(memo) else "Unknown"


st.set_page_config(page_title="FinTech Rec Engine", layout="wide")


# -----------------------------------------------------------------
# External CSS Loader
# -----------------------------------------------------------------
def load_external_css(css_file_path: str) -> None:
    try:
        with open(css_file_path, "r", encoding="utf-8") as css_file:
            st.markdown(f"<style>{css_file.read()}</style>", unsafe_allow_html=True)
    except FileNotFoundError:
        st.error(
            f"âš ï¸ Core Asset Missing: Could not find workspace sheet layout rules at '{css_file_path}'"
        )


# Load the enterprise corporate theme
load_external_css("style.css")


# -----------------------------------------------------------------
# Constants
# -----------------------------------------------------------------
PRIOR_PERIOD_EXCEPTION_COUNT = 71
PRIOR_PERIOD_EXCEPTION_VALUE = 243200.00

DEFAULT_METRICS_CACHE = {
    "count": 0,
    "val": 0.0,
    "pct": 0.0,
    "qb_rows": 0,
    "qb_vol": 0.0,
    "inf_rows": 0,
    "inf_vol": 0.0,
    "matched_rows": 0,
    "matched_vol": 0.0,
    "net_variance": 0.0,
    "critical_exception_count": 0,
    "high_exception_count": 0,
    "medium_exception_count": 0,
    "low_exception_count": 0,
    "health_score": 0,
    "health_status": "Not Run",
    "dominant_match_method": "N/A",
    "dominant_match_method_pct": 0.0,
}


# -----------------------------------------------------------------
# Session State Initialization
# -----------------------------------------------------------------
for key, default in {
    "processing_complete": False,
    "execution_time": 0.0,
    "last_run_time": "N/A",
    "selected_workbench_idx": None,
    "workbench_comments": {},
    "workbench_resolutions": {},
    "active_panel": "Dashboard Workspace",
    "net_variance": 0.0,
    "parsed_period": "05",
    "exceptions_df": pd.DataFrame(),
    "matched_df": pd.DataFrame(),
    "reconciled_qb_df": pd.DataFrame(),
    "agg_df": pd.DataFrame(),
    "match_method_df": pd.DataFrame(),
    "executive_summary_df": pd.DataFrame(),
    "mapped_columns": {},
    "metrics_cache": DEFAULT_METRICS_CACHE.copy(),
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

if "system_event_log" not in st.session_state:
    st.session_state.system_event_log = [
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] SYSTEM: Core AS400 Platform Initialized. Awaiting Data Ingestion Streams."
    ]


def append_system_event(event_message: str) -> None:
    st.session_state.system_event_log.append(
        f"[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}] {event_message}"
    )


# -----------------------------------------------------------------
# Analytics Helpers
# -----------------------------------------------------------------
def safe_currency(value) -> str:
    numeric_value = pd.to_numeric(value, errors="coerce")
    if pd.isna(numeric_value):
        numeric_value = 0.0
    return f"${numeric_value:,.2f}"


def classify_exception_materiality(amount) -> str:
    abs_amount = abs(pd.to_numeric(amount, errors="coerce"))
    if pd.isna(abs_amount):
        abs_amount = 0.0

    if abs_amount >= 10000:
        return "Critical"
    if abs_amount >= 1000:
        return "High"
    if abs_amount >= 100:
        return "Medium"
    return "Low"


def build_match_method_analytics(reconciled_df: pd.DataFrame, amount_col: str) -> pd.DataFrame:
    output_columns = [
        "Match Method",
        "Rows",
        "Matched Rows",
        "Exception Rows",
        "Gross Value",
        "Matched Value",
        "Exception Value",
        "Share of Rows",
        "Reliability Tier",
    ]

    if reconciled_df is None or reconciled_df.empty:
        return pd.DataFrame(columns=output_columns)

    df = reconciled_df.copy()

    if "IS_MATCHED" not in df.columns:
        df["IS_MATCHED"] = False

    df["IS_MATCHED"] = df["IS_MATCHED"].fillna(False).astype(bool)

    if "MATCH_METHOD" not in df.columns:
        df["MATCH_METHOD"] = df["IS_MATCHED"].map(
            {True: "Matched - Method Not Reported", False: "Unmatched / No Match"}
        )
    else:
        df["MATCH_METHOD"] = df["MATCH_METHOD"].fillna("").astype(str).str.strip()
        blank_method_mask = df["MATCH_METHOD"].eq("")
        df.loc[blank_method_mask, "MATCH_METHOD"] = df.loc[blank_method_mask, "IS_MATCHED"].map(
            {True: "Matched - Method Not Reported", False: "Unmatched / No Match"}
        )

    if amount_col not in df.columns:
        df["_MATCH_ANALYTICS_AMOUNT"] = 0.0
        amount_col = "_MATCH_ANALYTICS_AMOUNT"

    df["_AMOUNT_NUMERIC"] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    df["_ABS_AMOUNT"] = df["_AMOUNT_NUMERIC"].abs()
    df["_MATCHED_AMOUNT"] = df["_AMOUNT_NUMERIC"].where(df["IS_MATCHED"], 0.0)
    df["_EXCEPTION_AMOUNT"] = df["_AMOUNT_NUMERIC"].where(~df["IS_MATCHED"], 0.0)

    total_rows = len(df)

    grouped = (
        df.groupby("MATCH_METHOD", dropna=False)
        .agg(
            Rows=("MATCH_METHOD", "size"),
            Matched_Rows=("IS_MATCHED", "sum"),
            Gross_Value=("_ABS_AMOUNT", "sum"),
            Matched_Value=("_MATCHED_AMOUNT", "sum"),
            Exception_Value=("_EXCEPTION_AMOUNT", "sum"),
        )
        .reset_index()
    )

    grouped["Exception_Rows"] = grouped["Rows"] - grouped["Matched_Rows"]
    grouped["Share of Rows"] = grouped["Rows"] / total_rows if total_rows else 0.0

    def reliability_tier(method: str) -> str:
        method_upper = str(method).upper()
        if "PO" in method_upper and ("INV" in method_upper or "INVOICE" in method_upper):
            return "Strong"
        if "INVOICE" in method_upper or "INV" in method_upper:
            return "Moderate"
        if "FUZZY" in method_upper or "LEXICON" in method_upper or "MEMO" in method_upper:
            return "Review"
        if "UNMATCH" in method_upper or "NO MATCH" in method_upper:
            return "Exception"
        return "Standard"

    grouped["Reliability Tier"] = grouped["MATCH_METHOD"].apply(reliability_tier)
    grouped = grouped.sort_values(["Rows", "Gross_Value"], ascending=[False, False])

    grouped = grouped.rename(
        columns={
            "MATCH_METHOD": "Match Method",
            "Matched_Rows": "Matched Rows",
            "Exception_Rows": "Exception Rows",
            "Gross_Value": "Gross Value",
            "Matched_Value": "Matched Value",
            "Exception_Value": "Exception Value"
        }
    )

    return grouped[output_columns]


def calculate_health_score(match_pct: float, exception_value: float, qb_volume: float, net_variance: float) -> tuple[int, str]:
    if abs(pd.to_numeric(qb_volume, errors="coerce") or 0.0) < 0.01:
        return 0, "No Data"

    exposure_ratio = abs(exception_value) / abs(qb_volume)
    variance_ratio = abs(net_variance) / abs(qb_volume)

    match_component = max(min(match_pct, 100), 0) * 0.60
    exposure_component = max(100 - (exposure_ratio * 100), 0) * 0.25
    variance_component = max(100 - (variance_ratio * 100), 0) * 0.15

    score = round(match_component + exposure_component + variance_component)

    if score >= 95:
        status = "Excellent"
    elif score >= 85:
        status = "Acceptable"
    elif score >= 70:
        status = "Needs Review"
    else:
        status = "Critical Review"

    return score, status


def build_executive_summary_df(metrics: dict) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {"Metric": "QB Records", "Value": f"{metrics['qb_rows']:,}"},
            {"Metric": "AS400 Records", "Value": f"{metrics['inf_rows']:,}"},
            {"Metric": "Matched Records", "Value": f"{metrics['matched_rows']:,}"},
            {"Metric": "Open Exceptions", "Value": f"{metrics['count']:,}"},
            {"Metric": "Match Rate", "Value": f"{metrics['pct']:.1f}%"},
            {"Metric": "Open Exposure", "Value": safe_currency(metrics['val'])},
            {"Metric": "Net Variance", "Value": safe_currency(metrics['net_variance'])},
            {"Metric": "Health Score", "Value": f"{metrics['health_score']} / 100 - {metrics['health_status']}"},
            {"Metric": "Dominant Match Method", "Value": metrics["dominant_match_method"]},
        ]
    )


def render_summary_card(label: str, value: str, subtitle: str = "", use_navy: bool = False) -> None:
    card_class = "navy-kpi-card" if use_navy else "nested-metric-card"
    st.markdown(
        f"""
        <div class='{card_class}'>
            <span class='nested-metric-label'>{label}</span>
            <h2 class='nested-metric-value'>{value}</h2>
            <div class='nested-metric-subtitle'>{subtitle}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_executive_summary_dashboard(metrics: dict) -> None:
    st.markdown("<h5 style='color: var(--text-on-light); font-weight: 700; margin-top: 8px; margin-bottom: 4px;'>Executive Summary Dashboard</h5>", unsafe_allow_html=True)

    st.markdown("<div class='enterprise-card'>", unsafe_allow_html=True)
    row1 = st.columns(4)
    with row1[0]:
        render_summary_card("QB Records", f"{metrics['qb_rows']:,}", safe_currency(metrics["qb_vol"]), use_navy=True)
    with row1[1]:
        render_summary_card("AS400 Records", f"{metrics['inf_rows']:,}", safe_currency(metrics["inf_vol"]), use_navy=True)
    with row1[2]:
        render_summary_card("Matched Records", f"{metrics['matched_rows']:,}", safe_currency(metrics["matched_vol"]), use_navy=True)
    with row1[3]:
        render_summary_card("Open Exceptions", f"{metrics['count']:,}", safe_currency(metrics["val"]), use_navy=True)

    st.markdown("<div style='margin-top: 10px;'></div>", unsafe_allow_html=True)

    row2 = st.columns(4)
    with row2[0]:
        render_summary_card("Match Rate", f"{metrics['pct']:.1f}%", "Auto-reconciled records")
    with row2[1]:
        render_summary_card("Net Variance", safe_currency(metrics["net_variance"]), metrics["health_status"])
    with row2[2]:
        render_summary_card("Health Score", f"{metrics['health_score']} / 100", metrics["health_status"])
    with row2[3]:
        render_summary_card(
            "Dominant Match Method",
            f"{metrics['dominant_match_method_pct']:.1f}%",
            metrics["dominant_match_method"],
        )
    st.markdown("</div>", unsafe_allow_html=True)


def render_match_method_tracker(match_method_df: pd.DataFrame) -> None:
    st.markdown("<h5 style='color: var(--text-on-light); font-weight: 700; margin-top: 8px; margin-bottom: 4px;'>Match Method Analytics Tracker</h5>", unsafe_allow_html=True)

    if match_method_df.empty:
        st.info("No match method analytics are available until the reconciliation has run.")
        return

    st.markdown("<div class='enterprise-card'>", unsafe_allow_html=True)
    tracker_left, tracker_right = st.columns([1.4, 1])

    with tracker_left:
        display_df = match_method_df.copy()
        for currency_col in ["Gross Value", "Matched Value", "Exception Value"]:
            display_df[currency_col] = display_df[currency_col].apply(safe_currency)
        display_df["Share of Rows"] = display_df["Share of Rows"].apply(lambda x: f"{x:.1%}")
        st.dataframe(display_df, use_container_width=True, hide_index=True, height=170)

    with tracker_right:
        st.markdown(
            """
            <div style='background-color: var(--bg-light); padding: 10px; border-radius: 6px; border: 1px solid var(--border-light); height: 170px;'>
                <h5 style='color: var(--text-on-light); font-size:12px; font-weight:800; text-transform: uppercase; margin-top:0; margin-bottom: 4px;'>Method Distribution</h5>
            """,
            unsafe_allow_html=True,
        )
        
        # Native horizontal bar chart with primary corporate navy blue scheme
        top_methods = match_method_df.head(5).copy()
        if not top_methods.empty:
            chart_df = pd.DataFrame({
                "Volume (Rows)": top_methods["Rows"].values
            }, index=top_methods["Match Method"].values)
            st.bar_chart(chart_df, color="#1B365D", horizontal=True, height=120, use_container_width=True)
        else:
            st.caption("No method fields matched for visual distribution summaries.")

        st.markdown("</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------
# Sidebar Layout
# -----------------------------------------------------------------
st.sidebar.markdown("<h3 class='sidebar-header'>Workspace Routing</h3>", unsafe_allow_html=True)
st.sidebar.markdown("<p class='sidebar-subheader'>OPERATIONAL MODULES</p>", unsafe_allow_html=True)

current_panel = st.sidebar.radio(
    "Operational Workspace Modules",
    ["Dashboard Workspace", "Exception Workbench [Review]"],
    label_visibility="collapsed",
    index=["Dashboard Workspace", "Exception Workbench [Review]"].index(st.session_state.active_panel)
    if st.session_state.active_panel in ["Dashboard Workspace", "Exception Workbench [Review]"]
    else 0,
)
st.session_state.active_panel = current_panel

st.sidebar.markdown("<br><p class='sidebar-subheader'>ADMINISTRATION & CONTROLS</p>", unsafe_allow_html=True)
admin_panel = st.sidebar.radio(
    "Administrative Workspace Modules",
    ["System Audit Log", "Mapping Configuration"],
    label_visibility="collapsed",
)

if st.sidebar.button("ðŸ”“ Route to Selected Admin Panel", use_container_width=True):
    st.session_state.active_panel = admin_panel
    st.rerun()

current_panel = st.session_state.active_panel

with st.sidebar.expander("âš™ï¸ Schema Field Mapping", expanded=False):
    qb_po = st.text_input("QB PO Column", value="P.O. NUMBER")
    qb_inv = st.text_input("QB Invoice Column", value="NUM")
    qb_amt = st.text_input("QB Amount Column", value="AMOUNT")
    qb_qty = st.text_input("QB QTY Column", value="QTY")
    st.markdown("<hr style='border-top:1px solid #334155; margin:10px 0;'>", unsafe_allow_html=True)
    inf_po = st.text_input("Infinium PO Column", value="OHDESC")
    inf_inv = st.text_input("Infinium Invoice Column", value="OHOBNO")
    inf_amt = st.text_input("Infinium Amount Column", value="OHTOTA")

st.sidebar.markdown("<hr style='border-top:1px solid #1E293B; margin:15px 0;'>", unsafe_allow_html=True)
st.sidebar.markdown("<h4 class='sidebar-section-title'>ðŸ“¥ Data Source Feeds</h4>", unsafe_allow_html=True)

file_qb = st.sidebar.file_uploader("QuickBooks Export (.xlsx)", type=["xlsx"])
file_inf = st.sidebar.file_uploader("Infinium Data Pull (.xlsx)", type=["xlsx"])


# -----------------------------------------------------------------
# Source Data Extraction
# -----------------------------------------------------------------
qb_rows, qb_vol, inf_rows, inf_vol = 0, 0.0, 0, 0.0
raw_qb = None
df_qb_raw = pd.DataFrame()
df_inf_raw = pd.DataFrame()
df_qb_clean_rows = pd.DataFrame()
df_inf_clean_rows = pd.DataFrame()
qb_memo_col = None

if file_qb:
    raw_qb = pd.read_excel(file_qb, header=None)
    h_idx = 0

    for idx, row in raw_qb.head(15).iterrows():
        row_str_list = [str(val).strip().upper() for val in row.values if pd.notna(val)]
        if "AMOUNT" in row_str_list or "P.O. NUMBER" in row_str_list:
            h_idx = idx
            break

    df_qb_raw = pd.read_excel(file_qb, header=h_idx)
    df_qb_raw.columns = df_qb_raw.columns.astype(str).str.strip().str.upper()

    memo_candidates = [c for c in df_qb_raw.columns if "MEMO" in c]
    qb_memo_col = (
        "MEMO/DESCRIPTION"
        if "MEMO/DESCRIPTION" in df_qb_raw.columns
        else (memo_candidates[0] if memo_candidates else df_qb_raw.columns[0])
    )

    required_qb_cols = [qb_inv.upper(), qb_amt.upper(), qb_qty.upper(), qb_po.upper()]
    missing_qb_cols = [col for col in required_qb_cols if col not in df_qb_raw.columns]

    if missing_qb_cols:
        st.sidebar.error(f"QB missing required column(s): {', '.join(missing_qb_cols)}")
    else:
        df_qb_clean_rows = df_qb_raw[
            df_qb_raw[qb_inv.upper()].notna() | df_qb_raw[qb_memo_col].notna()
        ].dropna(how="all")
        qb_rows = len(df_qb_clean_rows)
        qb_vol = pd.to_numeric(df_qb_clean_rows[qb_amt.upper()], errors="coerce").sum()

if file_inf:
    df_inf_raw = pd.read_excel(file_inf)
    df_inf_raw.columns = df_inf_raw.columns.astype(str).str.strip().str.upper()

    required_inf_cols = [inf_po.upper(), inf_inv.upper(), inf_amt.upper()]
    missing_inf_cols = [col for col in required_inf_cols if col not in df_inf_raw.columns]

    if missing_inf_cols:
        st.sidebar.error(f"Infinium missing required column(s): {', '.join(missing_inf_cols)}")
    else:
        df_inf_clean_rows = df_inf_raw.dropna(how="all")
        inf_rows = len(df_inf_clean_rows)
        inf_vol = pd.to_numeric(df_inf_clean_rows[inf_amt.upper()], errors="coerce").sum()

if file_qb and file_inf and not df_qb_clean_rows.empty and not df_inf_clean_rows.empty:
    st.session_state.net_variance = qb_vol - inf_vol
    if raw_qb is not None and len(raw_qb) > 1 and pd.notna(raw_qb.iloc[1, 2]):
        st.session_state.parsed_period = str(raw_qb.iloc[1, 2]).strip()

    QB_COLUMNS = {
        "po": qb_po.upper(),
        "invoice": qb_inv.upper(),
        "amount": qb_amt.upper(),
        "qty": qb_qty.upper(),
    }
    AS400_COLUMNS = {
        "po": inf_po.upper(),
        "invoice": inf_inv.upper(),
        "amount": inf_amt.upper(),
    }
    st.session_state.mapped_columns = {"qb": QB_COLUMNS, "inf": AS400_COLUMNS, "qb_memo": qb_memo_col}


# -----------------------------------------------------------------
# Dashboard Workspace View
# -----------------------------------------------------------------
if current_panel == "Dashboard Workspace":
    
    # 1. Top Global Variance Notification Banner Strip
    if st.session_state.processing_complete:
        variance_label = "SYSTEM IN BALANCE" if round(st.session_state.net_variance, 2) == 0 else "SUBLEDGER OUT OF BALANCE"
        st.markdown(
            f"""
            <div class='global-alert-strip'>
                âš ï¸ {variance_label}: {safe_currency(st.session_state.net_variance)}
            </div>
            """,
            unsafe_allow_html=True,
        )

    # 2. Unified Control Panel Main Card
    st.markdown("<div class='enterprise-card' style='margin-top: 2px;'>", unsafe_allow_html=True)
    
    # Grid Header: Split Layout for Title Sizing Hierarchy & Form Button Action Placement
    head_col, action_col = st.columns([2.4, 1])
    
    with head_col:
        st.markdown(
            f"""
            <div style='margin-bottom: 6px;'>
                <h2 style='font-size: 18px; font-weight: 800; color: var(--text-on-light); margin: 0;'>Sales Reconciliation Command Center</h2>
                <div style='font-size: 11px; font-weight: 600; color: var(--text-on-light-muted); margin-top: 1px;'>
                    Period {st.session_state.parsed_period} Close &nbsp; â€¢ &nbsp; QuickBooks Online â†” Infinium AS400
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        
    with action_col:
        # Native form handlers sitting inside layout structure explicitly to avoid script execution drops
        btn_slots = st.columns([1.5, 0.9])
        with btn_slots[0]:
            run_triggered = st.button("Run Reconciliation", type="primary", use_container_width=True, disabled=(not file_qb or not file_inf))
        with btn_slots[1]:
            clear_triggered = st.button("Reset", type="secondary", use_container_width=True)

    # Refined Status Language and Dynamic Ice-Blue Focus Glow Classes
    if st.session_state.processing_complete:
        cls_step1 = "complete"
        cls_step2 = "complete"
        cls_step3 = "complete"
        cls_step4 = "active" 
        step3_str = "Complete"
        step4_str = "Review Required"
    else:
        cls_step1 = "complete" if file_qb else "active"
        cls_step2 = "complete" if file_inf else ("active" if file_qb else "")
        cls_step3 = "active" if (file_qb and file_inf) else "" 
        cls_step4 = ""
        step3_str = "Ready to Run" if (file_qb and file_inf) else "Pending Review"
        step4_str = "Pending Review"

    st.markdown(
        f"""
        <div class='process-flow-container'>
            <div class='flow-step {cls_step1}'>
                <b>1. Ingest QuickBooks</b><br><span>{"âœ”ï¸ Ingested" if file_qb else "Current Step"}</span>
            </div>
            <div class='flow-connector-line'></div>
            <div class='flow-step {cls_step2}'>
                <b>2. Ingest Infinium</b><br><span>{"âœ”ï¸ Ingested" if file_inf else ("Current Step" if file_qb else "Upcoming")}</span>
            </div>
            <div class='flow-connector-line'></div>
            <div class='flow-step {cls_step3}'>
                <b>3. Run Reconciliation</b><br><span>{step3_str}</span>
            </div>
            <div class='flow-connector-line'></div>
            <div class='flow-step {cls_step4}'>
                <b>4. Review Exceptions</b><br><span>{step4_str}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("</div>", unsafe_allow_html=True) # End Control Panel Card

    # Core Execution Form Arrays Ingest Handling Loops
    if run_triggered:
        if not st.session_state.mapped_columns:
            st.error("Schema mapping is not ready. Confirm both uploads contain the mapped columns.")
        else:
            start_timer = time.time()

            QB_COLUMNS = st.session_state.mapped_columns["qb"]
            AS400_COLUMNS = st.session_state.mapped_columns["inf"]
            qb_memo_col = st.session_state.mapped_columns["qb_memo"]

            df_qb_pass = df_qb_raw[
                df_qb_raw[QB_COLUMNS["invoice"]].notna() | df_qb_raw[qb_memo_col].notna()
            ].copy()
            df_qb_pass = df_qb_pass.dropna(how="all").reset_index(drop=True)
            df_inf_pass = df_inf_clean_rows.reset_index(drop=True)

            reconciled_qb, reconciled_inf = execute_sales_reconciliation(
                df_qb_pass, df_inf_pass, QB_COLUMNS, AS400_COLUMNS
            )

            if "IS_MATCHED" not in reconciled_qb.columns:
                reconciled_qb["IS_MATCHED"] = False
            reconciled_qb["IS_MATCHED"] = reconciled_qb["IS_MATCHED"].fillna(False).astype(bool)

            st.session_state.exceptions_df = reconciled_qb[~reconciled_qb["IS_MATCHED"]].copy()
            st.session_state.matched_df = reconciled_qb[reconciled_qb["IS_MATCHED"]].copy()
            st.session_state.reconciled_qb_df = reconciled_qb.copy()

            # Product aggregate sets
            df_qb_agg_source = df_qb_pass.copy()
            df_qb_agg_source[QB_COLUMNS["qty"]] = pd.to_numeric(
                df_qb_agg_source[QB_COLUMNS["qty"]], errors="coerce"
            ).fillna(0)
            df_qb_agg_source[QB_COLUMNS["amount"]] = pd.to_numeric(
                df_qb_agg_source[QB_COLUMNS["amount"]], errors="coerce"
            ).fillna(0)
            df_qb_agg_source["Product Name"] = df_qb_agg_source[qb_memo_col].apply(
                get_fuzzy_lexicon_match
            )
            df_qb_agg_filtered = df_qb_agg_source[df_qb_agg_source["Product Name"].notna()].copy()

            agg_df = (
                df_qb_agg_filtered.groupby("Product Name")
                .agg(
                    Product_Quantity=(QB_COLUMNS["qty"], "sum"),
                    Product_Value=(QB_COLUMNS["amount"], "sum"),
                )
                .reset_index()
            )

            st.session_state.agg_df = agg_df.sort_values(by="Product Name", ascending=True).rename(
                columns={"Product_Quantity": "Product Quantity", "Product_Value": "Product Value"}
            )

            total_scanned_qb = len(reconciled_qb)
            cleared_count_qb = len(st.session_state.matched_df)
            exception_count_qb = len(st.session_state.exceptions_df)
            pct_matched_qb = (cleared_count_qb / total_scanned_qb) * 100 if total_scanned_qb else 0.0
            exception_value_qb = pd.to_numeric(
                st.session_state.exceptions_df[QB_COLUMNS["amount"]], errors="coerce"
            ).sum()
            matched_value_qb = pd.to_numeric(
                st.session_state.matched_df[QB_COLUMNS["amount"]], errors="coerce"
            ).sum()

            exceptions_with_materiality = st.session_state.exceptions_df.copy()
            if not exceptions_with_materiality.empty and QB_COLUMNS["amount"] in exceptions_with_materiality.columns:
                exceptions_with_materiality["MATERIALITY_CLASS"] = exceptions_with_materiality[
                    QB_COLUMNS["amount"]
                ].apply(classify_exception_materiality)
                st.session_state.exceptions_df = exceptions_with_materiality

            materiality_counts = (
                st.session_state.exceptions_df["MATERIALITY_CLASS"].value_counts().to_dict()
                if "MATERIALITY_CLASS" in st.session_state.exceptions_df.columns
                else {}
            )

            match_method_df = build_match_method_analytics(reconciled_qb, QB_COLUMNS["amount"])
            st.session_state.match_method_df = match_method_df

            if not match_method_df.empty:
                dominant_row = match_method_df.iloc[0]
                dominant_match_method = str(dominant_row["Match Method"])
                dominant_match_method_pct = float(dominant_row["Share of Rows"] * 100)
            else:
                dominant_match_method = "N/A"
                dominant_match_method_pct = 0.0

            health_score, health_status = calculate_health_score(
                pct_matched_qb, exception_value_qb, qb_vol, qb_vol - inf_vol
            )

            st.session_state.metrics_cache = {
                "count": exception_count_qb,
                "val": exception_value_qb,
                "pct": pct_matched_qb,
                "qb_rows": qb_rows,
                "qb_vol": qb_vol,
                "inf_rows": inf_rows,
                "inf_vol": inf_vol,
                "matched_rows": cleared_count_qb,
                "matched_vol": matched_value_qb,
                "net_variance": qb_vol - inf_vol,
                "critical_exception_count": materiality_counts.get("Critical", 0),
                "high_exception_count": materiality_counts.get("High", 0),
                "medium_exception_count": materiality_counts.get("Medium", 0),
                "low_exception_count": materiality_counts.get("Low", 0),
                "health_score": health_score,
                "health_status": health_status,
                "dominant_match_method": dominant_match_method,
                "dominant_match_method_pct": dominant_match_method_pct,
            }

            st.session_state.executive_summary_df = build_executive_summary_df(
                st.session_state.metrics_cache
            )
            st.session_state.processing_complete = True
            st.session_state.execution_time = max(round(time.time() - start_timer, 2), 0.02)
            st.session_state.last_run_time = datetime.now().strftime("%I:%M %p")
            append_system_event(f"VALIDATION: Extracted {qb_rows} records from QBO workbook.")
            append_system_event(f"EXECUTION: Reconciliation loops compiled in {st.session_state.execution_time}s.")
            st.rerun()

    if clear_triggered:
        st.session_state.processing_complete = False
        st.session_state.execution_time = 0.0
        st.session_state.last_run_time = "N/A"
        st.session_state.selected_workbench_idx = None
        st.session_state.net_variance = 0.0
        st.session_state.exceptions_df = pd.DataFrame()
        st.session_state.matched_df = pd.DataFrame()
        st.session_state.reconciled_qb_df = pd.DataFrame()
        st.session_state.agg_df = pd.DataFrame()
        st.session_state.match_method_df = pd.DataFrame()
        st.session_state.executive_summary_df = pd.DataFrame()
        st.session_state.metrics_cache = DEFAULT_METRICS_CACHE.copy()
        st.rerun()

    # Ingested Ledger Pre-Execution Matrix Dashboard View
    if not st.session_state.processing_complete:
        if not file_qb or not file_inf:
            pass
        else:
            st.markdown("<div class='enterprise-card'>", unsafe_allow_html=True)
            st.markdown(
                """
                <div class='header-with-badge'>
                    <h4 style='color: var(--text-on-light); font-size:13px; font-weight:800; text-transform:uppercase; margin:0;'>Pre-Execution Ledger Verification</h4>
                    <span class='app-status-badge badge-validated'>Files Validated</span>
                </div>
                """, 
                unsafe_allow_html=True
            )
            
            st.markdown(
                f"""
                <div class='verification-kpi-grid'>
                    <div class='verification-tile'>
                        <div class='verification-tile-label'>QuickBooks Rows</div>
                        <div class='verification-tile-value'>{qb_rows:,}</div>
                    </div>
                    <div class='verification-tile'>
                        <div class='verification-tile-label'>Infinium Rows</div>
                        <div class='verification-tile-value'>{inf_rows:,}</div>
                    </div>
                    <div class='verification-tile alert'>
                        <div class='verification-tile-label'>Row Variance</div>
                        <div class='verification-tile-value'>{abs(qb_rows - inf_rows)}</div>
                    </div>
                    <div class='verification-tile alert'>
                        <div class='verification-tile-label'>Value Difference</div>
                        <div class='verification-tile-value'>{safe_currency(st.session_state.net_variance)}</div>
                    </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            
            st.markdown(
                f"""
                <table class='control-data-table'>
                    <thead>
                        <tr>
                            <th style='width: 45%;'>Source Ledger Stream</th>
                            <th style='text-align:right;'>Transaction Volume</th>
                            <th style='text-align:right;'>Control Gross Value</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>QuickBooks Online Subledger</td>
                            <td style='text-align:right;'>{qb_rows:,} rows</td>
                            <td style='text-align:right; font-weight:700;'>{safe_currency(qb_vol)}</td>
                        </tr>
                        <tr>
                            <td>Infinium AS400 Base Ledger</td>
                            <td style='text-align:right;'>{inf_rows:,} rows</td>
                            <td style='text-align:right; font-weight:700;'>{safe_currency(inf_vol)}</td>
                        </tr>
                        <tr style='background-color: var(--alert-bg); font-weight:700;'>
                            <td style='color: var(--alert-red); font-weight: 800;'>
                                Net Variance <span class='app-status-badge badge-out-of-balance' style='margin-left: 6px; padding: 1px 5px; font-size: 8px;'>Out of Balance</span>
                            </td>
                            <td style='text-align:right; color: var(--alert-red);'>{abs(qb_rows - inf_rows)} transactions | Infinium exceeds QuickBooks by {safe_currency(abs(st.session_state.net_variance))}</td>
                            <td style='text-align:right; color: var(--alert-red); font-weight: 800;'>{safe_currency(st.session_state.net_variance)}</td>
                        </tr>
                    </tbody>
                </table>
                <div style='margin-top: 8px; padding: 8px; background-color: #FFFBEB; border-radius: 4px; border: 1px solid #FDE68A; font-size: 12px; color: #B45309; font-weight:600;'>
                    ðŸ’¡ <b>Operational Conclusion:</b> Infinium exceeds QuickBooks by {safe_currency(abs(st.session_state.net_variance))} across {abs(qb_rows - inf_rows)} net rows. Verification adjustments required.
                </div>
                </div>
                """,
                unsafe_allow_html=True,
            )

    # Post-Execution Analytical Elements
    if st.session_state.processing_complete:
        c = st.session_state.metrics_cache

        render_executive_summary_dashboard(c)
        render_match_method_tracker(st.session_state.match_method_df)

        st.markdown("<div style='margin-top: 4px;'></div>", unsafe_allow_html=True)
        
        # Dual Segment Analytical Insight Grid
        st.markdown("<div class='enterprise-card'>", unsafe_allow_html=True)
        grid_col1, grid_col2 = st.columns(2)

        with grid_col1:
            exposure_change = (
                ((PRIOR_PERIOD_EXCEPTION_VALUE - c["val"]) / PRIOR_PERIOD_EXCEPTION_VALUE) * 100
                if PRIOR_PERIOD_EXCEPTION_VALUE
                else 0
            )

            st.markdown(
                f"""
                <div style='background-color: var(--bg-light); padding:10px; border-radius:6px; border:1px solid var(--border-light); height: 165px;'>
                    <h5 style='color: var(--text-on-light); font-size:12px; font-weight:800; text-transform: uppercase; margin-top:0;'>Closing Cycle Variance Trend</h5>
                    <table class='control-data-table' style='font-size:12px; margin-top:6px; color: var(--text-on-light) !important;'>
                        <thead>
                            <tr style='border-bottom: 2px solid var(--border-light); font-size:11px; font-weight:800;'>
                                <th style='padding:4px;'>Metric</th>
                                <th style='padding:4px; text-align:right;'>Prior P04</th>
                                <th style='padding:4px; text-align:right;'>Current P05</th>
                                <th style='padding:4px; text-align:right;'>Change</th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr style='border-bottom: 1px solid var(--border-light);'>
                                <td style='padding:6px 4px;'>Exceptions Count</td>
                                <td style='padding:6px 4px; text-align:right;'>{PRIOR_PERIOD_EXCEPTION_COUNT:,}</td>
                                <td style='padding:6px 4px; text-align:right;'>{c['count']:,}</td>
                                <td style='padding:6px 4px; text-align:right; color:#10B981; font-weight:700;'>
                                    {c['count'] - PRIOR_PERIOD_EXCEPTION_COUNT:+,}
                                </td>
                            </tr>
                            <tr style='border-bottom: 1px solid var(--border-light);'>
                                <td style='padding:6px 4px;'>Exposure Value</td>
                                <td style='padding:6px 4px; text-align:right;'>{safe_currency(PRIOR_PERIOD_EXCEPTION_VALUE)}</td>
                                <td style='padding:6px 4px; text-align:right;'>{safe_currency(c['val'])}</td>
                                <td style='padding:6px 4px; text-align:right; color:#10B981; font-weight:700;'>
                                    -{exposure_change:.1f}%
                                </td>
                            </tr>
                            <tr style='border-bottom: 1px solid var(--border-light);'>
                                <td style='padding:6px 4px;'>Critical Exceptions</td>
                                <td style='padding:6px 4px; text-align:right;'>N/A</td>
                                <td style='padding:6px 4px; text-align:right;'>{c['critical_exception_count']:,}</td>
                                <td style='padding:6px 4px; text-align:right; color: var(--text-on-light-muted);'>Current</td>
                            </tr>
                        </tbody>
                    </table>
                </div>
                """,
                unsafe_allow_html=True,
            )

        with grid_col2:
            st.markdown(
                """
                <div style='background-color: var(--bg-light); padding:10px; border-radius:6px; border:1px solid var(--border-light); height: 165px;'>
                    <h5 style='color: var(--text-on-light); font-size:12px; font-weight:800; text-transform: uppercase; margin-top:0; margin-bottom: 4px;'>Top Exception Drivers</h5>
                """,
                unsafe_allow_html=True,
            )

            driver_agg = st.session_state.agg_df.sort_values(by="Product Value", ascending=False).head(3)
            if driver_agg.empty:
                st.markdown(
                    "<p style='color: var(--text-on-light-muted); padding-top:10px; font-size:12px;'>No product aggregate drivers identified.</p>",
                    unsafe_allow_html=True,
                )
            else:
                max_driver_value = driver_agg["Product Value"].max()
                for _, d_row in driver_agg.iterrows():
                    bar_width = 0 if max_driver_value == 0 else (d_row["Product Value"] / max_driver_value) * 100
                    st.markdown(
                        f"""
                        <div class='driver-row-wrapper' style='margin-bottom: 4px;'>
                            <div class='driver-row-data' style='color: var(--text-on-light-secondary) !important; font-size: 11px;'>
                                <span>{d_row['Product Name']}</span>
                                <b>{safe_currency(d_row['Product Value'])}</b>
                            </div>
                            <div class='driver-bar-track' style='background: var(--border-light); height: 4px;'>
                                <div class='driver-bar-fill' style='width:{bar_width:.1f}%; background: var(--corporate-blue); height: 4px;'></div>
                            </div>
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )

            st.markdown("</div>", unsafe_allow_html=True)
        st.markdown("</div>", unsafe_allow_html=True)

        # -----------------------------------------------------------------
        # Excel Audit Package Export Engine
        # -----------------------------------------------------------------
        pkg_buffer = io.BytesIO()
        with pd.ExcelWriter(pkg_buffer, engine="openpyxl") as writer:
            st.session_state.executive_summary_df.to_excel(writer, index=False, sheet_name="0. Executive Summary")
            st.session_state.match_method_df.to_excel(writer, index=False, sheet_name="0. Match Method Analytics")
            st.session_state.exceptions_df.to_excel(writer, index=False, sheet_name="1. Unresolved Exceptions")
            st.session_state.matched_df.to_excel(writer, index=False, sheet_name="2. Matched Transactions")
            st.session_state.agg_df.to_excel(writer, index=False, sheet_name="3. Product Aggregates Summary")
            st.session_state.reconciled_qb_df.to_excel(writer, index=False, sheet_name="Raw QuickBooks Sales")

            workbook = writer.book
            workbook.calculation.calcMode = "auto"

            font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            font_body = Font(name="Segoe UI", size=11)
            font_total = Font(name="Segoe UI", size=11, bold=True)
            fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            fill_zebra = PatternFill(start_color="F4F7FA", end_color="F4F7FA", fill_type="solid")
            fill_total = PatternFill(start_color="EAECEF", end_color="EAECEF", fill_type="solid")
            fill_matched_grey = PatternFill(start_color="EBEBEB", end_color="EBEBEB", fill_type="solid")
            fill_exception_amber = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

            border_thin = Side(border_style="thin", color="D3D3D3")
            border_top_thick = Side(border_style="thin", color="000000")
            border_bottom_double = Side(border_style="double", color="000000")
            cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
            total_border = Border(top=border_top_thick, bottom=border_bottom_double)
            align_center = Alignment(horizontal="center", vertical="center")
            align_right = Alignment(horizontal="right", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")

            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                ws.views.sheetView[0].showGridLines = True
                ws.row_dimensions[1].height = 26

                for cell in ws[1]:
                    cell.font = font_header
                    cell.fill = fill_header
                    cell.alignment = align_center

                max_row = ws.max_row
                max_col = ws.max_column
                is_raw_sheet = sheet_name == "Raw QuickBooks Sales"
                match_flag_col_idx = None

                if is_raw_sheet:
                    headers = [str(ws.cell(row=1, column=c).value).upper() for c in range(1, max_col + 1)]
                    if "IS_MATCHED" in headers:
                        match_flag_col_idx = headers.index("IS_MATCHED") + 1

                for row_idx in range(2, max_row + 1):
                    ws.row_dimensions[row_idx].height = 20

                    if is_raw_sheet and match_flag_col_idx:
                        is_matched_val = ws.cell(row=row_idx, column=match_flag_col_idx).value
                        current_row_fill = (
                            fill_matched_grey
                            if str(is_matched_val).strip().upper() in ["TRUE", "1", "MATCHED"]
                            else fill_exception_amber
                        )
                    else:
                        current_row_fill = fill_zebra if row_idx % 2 == 0 else None

                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = font_body
                        cell.border = cell_border

                        if current_row_fill:
                            cell.fill = current_row_fill

                        header_text = str(ws.cell(row=1, column=col_idx).value).upper()

                        if "AMOUNT" in header_text or "VALUE" in header_text or "OHTOTA" in header_text:
                            cell.number_format = "$#,##0.00"
                            cell.alignment = align_right
                        elif "QTY" in header_text or "QUANTITY" in header_text or "ROWS" in header_text:
                            cell.number_format = "#,##0"
                            cell.alignment = align_right
                        elif "PCT" in header_text or "SHARE" in header_text or "RATE" in header_text:
                            cell.number_format = "0.0%"
                            cell.alignment = align_right
                        elif "DATE" in header_text or "IS_MATCHED" in header_text:
                            cell.alignment = align_center
                        else:
                            cell.alignment = align_left

                if max_row > 1 and sheet_name not in ["Raw QuickBooks Sales", "0. Executive Summary", "0. Match Method Analytics"]:
                    tot_row_idx = max_row + 1
                    ws.row_dimensions[tot_row_idx].height = 22
                    ws.cell(row=tot_row_idx, column=1, value="Total").font = font_total

                    for col_idx in range(1, max_col + 1):
                        t_cell = ws.cell(row=tot_row_idx, column=col_idx)
                        t_cell.fill = fill_total
                        t_cell.border = total_border
                        t_cell.font = font_total

                        h_text = str(ws.cell(row=1, column=col_idx).value).upper()
                        col_letter = get_column_letter(col_idx)

                        if col_idx > 1 and ("AMOUNT" in h_text or "VALUE" in h_text or col_letter in ["G", "H"]):
                            t_cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
                            t_cell.number_format = "$#,##0.00"
                            t_cell.alignment = align_right
                        elif col_idx > 1 and ("QTY" in h_text or "QUANTITY" in h_text):
                            t_cell.value = f"=SUM({col_letter}2:{col_letter}{max_row})"
                            t_cell.number_format = "#,##0"
                            t_cell.alignment = align_right

                for col in ws.columns:
                    max_len = 0
                    col_letter = get_column_letter(col[0].column)

                    for cell in col:
                        if cell.value is not None:
                            val_str = str(cell.value)
                            if cell.number_format and "$" in cell.number_format:
                                val_str += "   "
                            max_len = max(max_len, len(val_str))

                    ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

        st.markdown("<h5 style='color: var(--text-on-light); font-weight:700; margin-top:4px;'>System Package Actions</h5>", unsafe_allow_html=True)
        st.markdown("<div class='enterprise-card' style='padding: 6px; margin-top: 4px;'>", unsafe_allow_html=True)
        if st.download_button(
            label="ðŸ“¥ Generate Institutional Audit Package (.XLSX)",
            data=pkg_buffer.getvalue(),
            file_name="Sales_Reconciliation_Master_Package.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        ):
            append_system_event("EXPORTS: Unified corporate excel package compiled cleanly.")
        st.markdown("</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------
# Exception Workbench View
# -----------------------------------------------------------------
elif current_panel == "Exception Workbench [Review]":
    if not st.session_state.processing_complete or st.session_state.exceptions_df.empty:
        st.warning("âš ï¸ Access Denied: Execute close processes on the dashboard panel to load open subledger items.")
    else:
        st.markdown("<h3 style='color: var(--text-on-light); font-weight:700;'>ðŸ” Exception Review Workbench</h3>", unsafe_allow_html=True)

        st.markdown("<div class='enterprise-card'>", unsafe_allow_html=True)
        list_col, work_col = st.columns([2, 2])
        QB_COLUMNS = st.session_state.mapped_columns["qb"]
        exceptions_df_raw = st.session_state.exceptions_df

        with list_col:
            st.markdown("#### Open Exceptions Vector")

            selection_options = [
                f"Row {idx} | PO: {row.get(QB_COLUMNS['po'], 'N/A')} | Amt: {safe_currency(row.get(QB_COLUMNS['amount'], 0))}"
                for idx, row in exceptions_df_raw.iterrows()
            ]

            if not selection_options:
                st.success("ðŸŽ‰ Zero exceptions detected. All ledger lines cleared autonomously.")
            else:
                default_sel_idx = 0
                if st.session_state.selected_workbench_idx in exceptions_df_raw.index:
                    current_idx_str = f"Row {st.session_state.selected_workbench_idx} "
                    matching_opt = [o for o in selection_options if o.startswith(current_idx_str)]
                    if matching_opt:
                        default_sel_idx = selection_options.index(matching_opt[0])

                selected_item_str = st.selectbox(
                    "Select Variance Row Target",
                    selection_options,
                    label_visibility="collapsed",
                    index=default_sel_idx,
                )

                selected_idx = int(selected_item_str.split(" | ")[0].replace("Row ", ""))
                st.session_state.selected_workbench_idx = selected_idx

                st.markdown("<div class='quick-entry-note-title'><b>ðŸ“ Quick Entry Item Note:</b></div>", unsafe_allow_html=True)

                existing_quick_note = st.session_state.workbench_comments.get(selected_idx, "")
                updated_note = st.text_input(
                    "Add / Edit clearance details for selected row",
                    value=existing_quick_note,
                    key=f"quick_note_{selected_idx}",
                    label_visibility="collapsed",
                    placeholder="e.g., AP is researching, received via wire, transit variance...",
                )

                if updated_note != existing_quick_note:
                    st.session_state.workbench_comments[selected_idx] = updated_note
                    append_system_event(f"NOTES: Captured audit text statement on index row #{selected_idx}.")

                st.dataframe(exceptions_df_raw, use_container_width=True)

        with work_col:
            if (
                st.session_state.selected_workbench_idx is not None
                and st.session_state.selected_workbench_idx in exceptions_df_raw.index
            ):
                target_row = exceptions_df_raw.loc[st.session_state.selected_workbench_idx]

                ranked_exceptions = exceptions_df_raw.copy()
                ranked_exceptions["_ABS_AMOUNT"] = pd.to_numeric(
                    ranked_exceptions[QB_COLUMNS["amount"]], errors="coerce"
                ).abs()

                ranked_exceptions["_RANK_BY_VALUE"] = ranked_exceptions["_ABS_AMOUNT"].rank(
                    method="dense", ascending=False
                ).fillna(1).astype(int)

                target_abs_amount = abs(pd.to_numeric(target_row.get(QB_COLUMNS["amount"], 0), errors="coerce"))
                target_rank = int(
                    ranked_exceptions.loc[st.session_state.selected_workbench_idx, "_RANK_BY_VALUE"]
                )

                exception_value_qb = st.session_state.metrics_cache["val"]
                target_exposure_pct = (
                    target_abs_amount / abs(exception_value_qb) * 100 if exception_value_qb else 0
                )

                st.markdown(f"### Mitigate Row Entry Reference #{st.session_state.selected_workbench_idx}")

                m_col1, m_col2, m_col3 = st.columns(3)
                m_col1.metric("Total Open Exceptions", f"{st.session_state.metrics_cache['count']}")
                m_col2.metric("Exposure Value Rank", f"#{target_rank}")
                m_col3.metric("Total Exposure Contribution", f"{target_exposure_pct:.2f}%")

                st.markdown("---")

                po_num_val = target_row.get(QB_COLUMNS["po"], "N/A")
                inv_num_val = target_row.get(QB_COLUMNS["invoice"], "N/A")
                raw_amt_val = target_row.get(QB_COLUMNS["amount"], 0.0)
                materiality_val = target_row.get("MATERIALITY_CLASS", classify_exception_materiality(raw_amt_val))
                match_method_val = target_row.get("MATCH_METHOD", "Unmatched / No Match")

                st.markdown(f"**Purchase Order Number:** `{po_num_val}`")
                st.markdown(f"**Invoice Number:** `{inv_num_val}`")
                st.markdown(f"**QuickBooks Base Amount:** `{safe_currency(raw_amt_val)}`")
                st.markdown(f"**Materiality Class:** :red[{materiality_val}]")
                st.markdown(f"**Match Method:** `{match_method_val}`")
                st.markdown("**Exception Classification:** :red[Missing AS400 Transaction]")
                st.markdown(f"**Granular Item Variance:** :red[{safe_currency(raw_amt_val)}]")

                st.markdown("---")
                st.markdown("##### Resolution Actions Management")

                current_saved_res = st.session_state.workbench_resolutions.get(
                    st.session_state.selected_workbench_idx, "Unassigned Open Item"
                )
                current_saved_comment = st.session_state.workbench_comments.get(
                    st.session_state.selected_workbench_idx, ""
                )

                resolution_options = [
                    "Unassigned Open Item",
                    "Timing Difference - Transit Accrual Required",
                    "Freight Classification Variance Adjustment",
                    "Approved AS400 Variance Write-off Entry",
                ]

                resolution_type = st.selectbox(
                    "Assign Action Status",
                    resolution_options,
                    index=resolution_options.index(current_saved_res)
                    if current_saved_res in resolution_options
                    else 0,
                )

                audit_comment = st.text_area(
                    "Audit Log Clearance Comments (Mirrored)",
                    value=current_saved_comment,
                    key=f"detailed_note_{st.session_state.selected_workbench_idx}",
                )

                if audit_comment != current_saved_comment:
                    st.session_state.workbench_comments[st.session_state.selected_workbench_idx] = audit_comment

                w_btn1, w_btn2 = st.columns(2)

                with w_btn1:
                    if st.button("Commit Resolution State", type="primary", use_container_width=True):
                        st.session_state.workbench_resolutions[
                            st.session_state.selected_workbench_idx
                        ] = resolution_type
                        append_system_event(
                            f"MITIGATION: Override status locked on index row #{st.session_state.selected_workbench_idx} [Status: {resolution_type}]."
                        )
                        st.success(f"âœ“ Resolution assigned to row #{st.session_state.selected_workbench_idx}.")

                with w_btn2:
                    if st.button("Clear Selection", use_container_width=True):
                        st.session_state.selected_workbench_idx = None
                        st.rerun()

                if st.session_state.selected_workbench_idx in st.session_state.workbench_resolutions:
                    st.markdown(
                        f"""
                        <div class='resolution-profile-lockbox'>
                            <b>ðŸ”’ Documented Resolution Profile:</b><br>
                            â€¢ Status Type: {st.session_state.workbench_resolutions[st.session_state.selected_workbench_idx]}<br>
                            â€¢ Comment: "{st.session_state.workbench_comments.get(st.session_state.selected_workbench_idx, '')}"
                        </div>
                        """,
                        unsafe_allow_html=True,
                    )
        st.markdown("</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------
# System Audit Log View
# -----------------------------------------------------------------
elif current_panel == "System Audit Log":
    st.markdown("<h3 style='color: var(--text-on-light); font-weight:700;'>ðŸ“‹ AS400 System Event Log & Security Audit Trail</h3>", unsafe_allow_html=True)
    st.markdown(
        "<p style='color:#64748B; font-size:14px; margin-bottom:20px;'>SOC 1/2 Compliance Stream: Verifiable platform orchestration and system matching execution logs:</p>",
        unsafe_allow_html=True,
    )
    
    st.markdown("<div class='system-log-container'>", unsafe_allow_html=True)
    for log_record in st.session_state.system_event_log:
        st.markdown(f"<div class='system-log-line'>{log_record}</div>", unsafe_allow_html=True)
    st.markdown("</div>", unsafe_allow_html=True)


# -----------------------------------------------------------------
# Mapping Configuration View
# -----------------------------------------------------------------
elif current_panel == "Mapping Configuration":
    st.markdown("<h3 style='color: var(--text-on-light); font-weight:700;'>âš™ï¸ Mapping Configuration Settings</h3>", unsafe_allow_html=True)
    st.markdown(
        """
        <div class='enterprise-card'>
            <h5 style='color: var(--corporate-blue); margin-top:0; font-weight:800; text-transform:uppercase;'>ðŸ”’ Environment Execution Constants</h5>
            <table class='control-data-table' style='font-size:13px; color: var(--text-on-light) !important;'>
                <tr style='border-bottom: 1px solid var(--border-light);'><td style='padding:12px 0; font-weight:600; width:40%;'>System Hosting Stack:</td><td>Local Port Server Endpoint</td></tr>
                <tr style='border-bottom: 1px solid var(--border-light);'><td style='padding:12px 0; font-weight:600;'>Array Engine Driver:</td><td>Multi-Pass Dataframe Matching Engine</td></tr>
                <tr style='border-bottom: 1px solid var(--border-light);'><td style='padding:12px 0; font-weight:600;'>Reconciliation Security State:</td><td>Stable Session Container Confirmed</td></tr>
                <tr style='border-bottom: 1px solid var(--border-light);'><td style='padding:12px 0; font-weight:600;'>Dashboard Enhancements:</td><td>Executive Summary + Match Method Analytics Enabled</td></tr>
            </table>
        </div>
        """,
        unsafe_allow_html=True,
    )
